# -*- coding: utf-8 -*-
"""
Імпорт оцінок з Excel ("зведена відомість").

Майстер з 4 кроків:
  1. Завантаження файлу + вибір групи
  2. Вибір аркуша (якщо їх декілька)
  3. Зіставлення: колонки-предмети з Excel <-> предмети групи в базі,
     рядки-студенти з Excel <-> студенти групи (автоматичне нечітке
     порівняння з можливістю поправити вручну)
  4. Попередній перегляд і підтвердження

Підтримує обидва реальні формати відомостей:
  А) "зведена екзаменаційна відомість": предмети в шапці, під кожним -
     3 колонки (Бали / Оцінка / ECTS, порядок буває переплутаний, тому
     значення визначається за ВМІСТОМ - береться число 0..100);
  Б) простий лист: один стовпець на предмет, у клітинках одразу бали.

Стан між кроками зберігається у тимчасових json-файлах на сервері
(папка import_sessions/), а не в cookie-сесії - розібраний Excel може
бути завеликим для cookie.
"""
import os
import re
import json
import time
import uuid
import difflib

import openpyxl
from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from werkzeug.utils import secure_filename

from routes.db import get_db
from routes.utils import permission_required, log_action, logger
from routes.helpers import current_username

import_grades_bp = Blueprint('import_grades', __name__)

SESSIONS_DIR = os.path.join(os.getcwd(), 'import_sessions')
SESSION_TTL_SECONDS = 6 * 3600


# ============================================================
# Допоміжне: тимчасові файли стану майстра
# ============================================================

def _cleanup_old_sessions():
    if not os.path.isdir(SESSIONS_DIR):
        return
    now = time.time()
    for f in os.listdir(SESSIONS_DIR):
        path = os.path.join(SESSIONS_DIR, f)
        try:
            if os.path.isfile(path) and now - os.path.getmtime(path) > SESSION_TTL_SECONDS:
                os.remove(path)
        except OSError:
            pass


def _state_path(token):
    return os.path.join(SESSIONS_DIR, f"{token}.json")


def _xlsx_path(token):
    return os.path.join(SESSIONS_DIR, f"{token}.xlsx")


def _save_state(token, state):
    os.makedirs(SESSIONS_DIR, exist_ok=True)
    with open(_state_path(token), 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False)


def _load_state_or_none(token):
    token = re.sub(r'[^0-9a-f]', '', token or '')
    path = _state_path(token)
    if not token or not os.path.isfile(path):
        return None
    with open(path, encoding='utf-8') as f:
        return json.load(f)


# ============================================================
# Нормалізація і нечітке порівняння
# ============================================================

def _norm_subject(s):
    """Нормалізує назву предмета для порівняння: без регістру, без
    приміток типу "(іспит)"/"(залік)", без пунктуації і зайвих пробілів."""
    s = str(s or '').lower()
    s = re.sub(r'\((?:[^)]*?(?:іспит|испит|залік|зачет|екзамен|экзамен)[^)]*?)\)', ' ', s)
    s = re.sub(r'[^\w\s]', ' ', s, flags=re.UNICODE)
    return ' '.join(s.split())


def _norm_name(s):
    """Нормалізує ПІБ: без дужок (дівоче прізвище), регістру, пунктуації."""
    s = str(s or '')
    s = re.sub(r'\(.*?\)', ' ', s)
    s = s.replace('ʼ', "'").replace('’', "'")
    s = re.sub(r"[^\w\s'-]", ' ', s.lower(), flags=re.UNICODE)
    return ' '.join(s.split())


def _ratio(a, b):
    return difflib.SequenceMatcher(None, a, b).ratio()


def _match_name(excel_name, db_students):
    """
    Повертає (student_id | None, score, status) для рядка з Excel.
    db_students: список dict {'id', 'full_name'} (ПІБ у порядку
    Прізвище Ім'я По-батькові, як зберігається в базі).
    Обробляє і повні ПІБ, і скорочення з ініціалами ("Прізвище І.П.").
    """
    raw = str(excel_name or '').strip()
    norm = _norm_name(raw)
    if not norm:
        return None, 0.0, 'none'

    has_initials = bool(re.search(r'\b[а-яіїєґa-z]\s*[.]', raw.lower()))

    best_id, best_score = None, 0.0
    for st in db_students:
        cand = _norm_name(st['full_name'])
        if has_initials:
            # Порівнюємо прізвище + ініціали
            excel_parts = norm.replace('.', ' ').split()
            cand_parts = cand.split()
            if not excel_parts or not cand_parts:
                continue
            surname_score = _ratio(excel_parts[0], cand_parts[0])
            initials_excel = [p[0] for p in excel_parts[1:3]]
            initials_cand = [p[0] for p in cand_parts[1:3]]
            init_ok = initials_excel == initials_cand[:len(initials_excel)]
            score = surname_score * (1.0 if init_ok else 0.6)
        else:
            score = _ratio(norm, cand)
        if score > best_score:
            best_id, best_score = st['id'], score

    if best_score >= 0.87:
        return best_id, best_score, 'auto'
    if best_score >= 0.65:
        return best_id, best_score, 'suggest'
    return None, best_score, 'none'


def _match_subject(excel_title, db_subjects):
    """Повертає (subject_id | None, score, status)."""
    norm = _norm_subject(excel_title)
    if not norm:
        return None, 0.0, 'none'
    best_id, best_score = None, 0.0
    for sub in db_subjects:
        score = _ratio(norm, _norm_subject(sub['name']))
        if score > best_score:
            best_id, best_score = sub['id'], score
    if best_score >= 0.85:
        return best_id, best_score, 'auto'
    if best_score >= 0.60:
        return best_id, best_score, 'suggest'
    return None, best_score, 'none'


# ============================================================
# Розбір аркуша Excel
# ============================================================

_SERVICE_TITLES = ('№', 'п/п', 'з/п', 'прізвище', 'є заяви', 'бали', 'оцінка', 'ects', 'семестр')


def _is_service_title(text):
    t = str(text).strip().lower()
    return (not t) or any(t.startswith(s) or s in t for s in _SERVICE_TITLES) or len(t) < 3


def _has_cyrillic_words(text, min_words=2):
    words = re.findall(r'[А-ЯІЇЄҐа-яіїєґ][а-яіїєґА-ЯІЇЄҐ\'ʼ’-]{1,}', str(text or ''))
    return len(words) >= min_words


def _cell_grade_value(v):
    """Число 0..100 з клітинки (int/float/рядок з числом) або None."""
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        n = int(v)
        return n if 0 <= n <= 100 and float(v) == n else None
    m = re.fullmatch(r'\s*(\d{1,3})\s*', str(v))
    if m:
        n = int(m.group(1))
        return n if 0 <= n <= 100 else None
    return None


def parse_sheet(ws):
    """
    Розбирає аркуш у структуру:
      {'subjects': [{'title', 'col'}],           # колонки-предмети
       'students': [{'name', 'row'}],            # рядки-студенти
       'grades': {'row,col': бал}}               # знайдені числа
    Автоматично розпізнає формат А (3-колонкові блоки з ECTS) та
    формат Б (одна колонка на предмет).
    """
    max_row, max_col = ws.max_row, ws.max_column
    grid = [[ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
            for r in range(1, max_row + 1)]

    # ---- Формат А: шукаємо рядок із кількома "ECTS" ----
    subheader_r = None
    for r, row in enumerate(grid):
        ects_count = sum(1 for v in row if isinstance(v, str) and v.strip().upper() == 'ECTS')
        if ects_count >= 2:
            subheader_r = r
            break

    subjects, name_col, data_start = [], None, None

    if subheader_r is not None and subheader_r >= 2:
        subject_r = subheader_r - 2  # рядок назв (між ними - рядок кредитів)
        data_start = subheader_r + 1

        titles = [(c, str(v).strip()) for c, v in enumerate(grid[subject_r])
                  if v is not None and str(v).strip() and not _is_service_title(v)]
        # Межі блоку предмета: від його колонки до колонки наступного предмета
        for i, (c, title) in enumerate(titles):
            end = titles[i + 1][0] if i + 1 < len(titles) else max_col
            subjects.append({'title': title, 'col': c, 'col_end': end})

        first_subj_col = titles[0][0] if titles else 3
        # Колонка ПІБ: серед службових колонок зліва — та, де найбільше кирилічних імен
        best_c, best_hits = None, 0
        for c in range(0, first_subj_col):
            hits = sum(1 for r in range(data_start, max_row)
                       if _has_cyrillic_words(grid[r][c] if c < len(grid[r]) else None))
            if hits > best_hits:
                best_c, best_hits = c, hits
        name_col = best_c if best_c is not None else 2

    else:
        # ---- Формат Б: рядок назв + одна колонка на предмет ----
        # Межа пошуку шапки - навмисно велика (не 5-6 рядків): у реальних
        # файлах перед таблицею часто йде багаторядковий титул/назва
        # спеціальності/порожні рядки (зустрічався файл, де шапка
        # предметів була аж на 7-му рядку - при межі "перші 6 рядків"
        # вона просто не потрапляла в діапазон пошуку).
        subject_r = None
        header_search_limit = min(40, max_row - 1)
        for r in range(0, header_search_limit):
            row = grid[r]
            text_cells = [(c, str(v).strip()) for c, v in enumerate(row)
                          if isinstance(v, str) and len(str(v).strip()) >= 4 and not _is_service_title(v)]
            if len(text_cells) >= 3:
                # наступний непорожній рядок має містити числа
                for rr in range(r + 1, min(r + 4, max_row)):
                    nums = sum(1 for v in grid[rr] if _cell_grade_value(v) is not None)
                    if nums >= 2:
                        subject_r = r
                        data_start = rr
                        break
            if subject_r is not None:
                break
        if subject_r is None:
            return {'subjects': [], 'students': [], 'grades': {}}

        # Колонка ПІБ: та, де в рядках даних кирилічні імена
        best_c, best_hits = 0, 0
        for c in range(0, max_col):
            hits = sum(1 for r in range(data_start, max_row)
                       if _has_cyrillic_words(grid[r][c] if c < len(grid[r]) else None))
            if hits > best_hits:
                best_c, best_hits = c, hits
        name_col = best_c

        for c, v in enumerate(grid[subject_r]):
            if c == name_col or v is None:
                continue
            title = str(v).strip()
            if title and not _is_service_title(title):
                subjects.append({'title': title, 'col': c, 'col_end': c + 1})

    # ---- Студенти і оцінки ----
    # Зупиняємось на першому повністю порожньому рядку - це природна
    # межа кінця таблиці. Без цього текст ПІСЛЯ таблиці (підсумки,
    # підписи на кшталт "Декан факультету: Прізвище Ім'я" в тій самій
    # колонці, що й ПІБ студентів) міг би сприйматися як додаткові
    # фальшиві "студенти".
    students, grades = [], {}
    for r in range(data_start, max_row):
        row = grid[r]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            break

        name_val = row[name_col] if name_col < len(row) else None
        if not _has_cyrillic_words(name_val):
            continue
        students.append({'name': str(name_val).strip(), 'row': r})
        for si, sub in enumerate(subjects):
            val = None
            for c in range(sub['col'], sub['col_end']):
                if c < len(grid[r]):
                    val = _cell_grade_value(grid[r][c])
                    if val is not None:
                        break
            if val is not None:
                grades[f"{len(students)-1},{si}"] = val

    return {'subjects': [{'title': s['title']} for s in subjects],
            'students': [{'name': s['name']} for s in students],
            'grades': grades}


def validate_parsed(parsed):
    """
    Перевіряє результат parse_sheet() на ознаки того, що структура файлу
    насправді не відповідає жодному з підтримуваних форматів як слід -
    навіть якщо технічно щось знайшлося (предмети/студенти не порожні).
    Повертає список текстових попереджень (не помилок - імпорт все одно
    може тривати далі, це підказки для уважнішої перевірки на кроках
    зіставлення й попереднього перегляду).
    """
    warnings = []
    subjects = parsed['subjects']
    students = parsed['students']
    grades = parsed['grades']

    if not subjects or not students:
        return warnings  # порожній результат - окреме повідомлення в choose_sheet()

    total_cells = len(subjects) * len(students)
    fill_ratio = len(grades) / total_cells if total_cells else 0

    # 1. Дуже мало реально заповнених оцінок відносно розміру таблиці -
    #    ознака того, що предмети чи студенти визначені неправильно
    #    (напр. невдало обрана колонка ПІБ, чи межі блоку предмета).
    if fill_ratio < 0.3:
        warnings.append(
            f"Заповнено лише {round(fill_ratio*100)}% очікуваних оцінок "
            f"({len(grades)} з {total_cells}) - можливо, предмети або студенти "
            f"визначені неправильно. Уважно перевірте крок зіставлення."
        )

    # 2. Шкала оцінок - якщо майже всі знайдені значення дуже малі,
    #    ймовірно це не 100-бальна шкала ECTS (наприклад, 5-бальна чи
    #    12-бальна), а парсер підставить ці числа як є, спотворивши сенс
    #    ("5" стане "Незадовільно" за шкалою 0-100).
    values = list(grades.values())
    if values:
        max_val = max(values)
        share_low = sum(1 for v in values if v <= 12) / len(values)
        if max_val <= 12 and share_low > 0.8:
            warnings.append(
                f"Знайдені оцінки переважно дуже малі (максимум серед знайдених: {max_val}) - "
                f"схоже, це не 100-бальна шкала ECTS, а інша система оцінювання. "
                f"Система інтерпретує ці числа як бали зі шкали 0-100 - "
                f"перевірте перед підтвердженням, інакше оцінки будуть спотворені."
            )

    # 3. Лише 1 предмет знайдено у файлі, схожому на формат А (є хоча б
    #    одне слово ECTS у файлі, але недостатньо для впевненого
    #    визначення 3-колонкового формату) - типова ознака помилкового
    #    "провалювання" у формат Б.
    if len(subjects) == 1:
        warnings.append(
            "Знайдено лише 1 предмет. Якщо у файлі їх насправді більше - "
            "можливо, структуру розпізнано неправильно (наприклад, формат "
            "із колонками Бали/Оцінка/ECTS визначається впевнено лише коли "
            "предметів у файлі щонайменше 2). Перевірте результат уважно."
        )

    return warnings


# ============================================================
# Маршрути майстра
# ============================================================

@import_grades_bp.route('/admin/import_grades', methods=['GET', 'POST'])
@permission_required('import_grades')
def upload():
    """Крок 1: файл + група."""
    _cleanup_old_sessions()
    conn = get_db()
    groups = conn.execute(
        "SELECT id, name, start_year, study_form FROM groups WHERE COALESCE(archived, 0) NOT IN (1) ORDER BY name"
    ).fetchall()
    conn.close()

    if request.method == 'POST':
        file = request.files.get('excel_file')
        group_id = request.form.get('group_id', type=int)
        if not file or not file.filename.lower().endswith('.xlsx'):
            flash('Оберіть файл формату .xlsx', 'danger')
            return redirect(url_for('import_grades.upload'))
        if not group_id:
            flash('Оберіть групу', 'danger')
            return redirect(url_for('import_grades.upload'))

        token = uuid.uuid4().hex
        os.makedirs(SESSIONS_DIR, exist_ok=True)
        file.save(_xlsx_path(token))

        try:
            wb = openpyxl.load_workbook(_xlsx_path(token), data_only=True, read_only=True)
            sheets = wb.sheetnames
            wb.close()
        except Exception as e:
            logger.error(f"Імпорт оцінок: не вдалося відкрити файл: {e}", exc_info=True)
            flash(f'Не вдалося відкрити файл: {e}', 'danger')
            return redirect(url_for('import_grades.upload'))

        _save_state(token, {
            'group_id': group_id,
            'filename': secure_filename(file.filename),
            'sheets': sheets,
            'user_id': session.get('user_id'),
        })
        return redirect(url_for('import_grades.choose_sheet', token=token))

    return render_template('import_grades.html', groups=groups)


@import_grades_bp.route('/admin/import_grades/<token>/sheet', methods=['GET', 'POST'])
@permission_required('import_grades')
def choose_sheet(token):
    """Крок 2: вибір аркуша (одразу переходить далі, якщо аркуш один)."""
    state = _load_state_or_none(token)
    if not state:
        flash('Сесію імпорту не знайдено або вона застаріла - завантажте файл ще раз', 'warning')
        return redirect(url_for('import_grades.upload'))

    if request.method == 'POST' or len(state['sheets']) == 1:
        sheet = request.form.get('sheet') or state['sheets'][0]
        if sheet not in state['sheets']:
            flash('Невідомий аркуш', 'danger')
            return redirect(url_for('import_grades.choose_sheet', token=token))

        wb = openpyxl.load_workbook(_xlsx_path(token), data_only=True)
        parsed = parse_sheet(wb[sheet])
        wb.close()

        if not parsed['subjects'] or not parsed['students']:
            flash('Не вдалося розпізнати структуру аркуша: не знайдено предметів або студентів. '
                  'Перевірте, що обрано правильний аркуш зведеної відомості.', 'danger')
            return redirect(url_for('import_grades.choose_sheet', token=token))

        for warning_text in validate_parsed(parsed):
            flash(warning_text, 'warning')

        state['sheet'] = sheet
        state['parsed'] = parsed
        _save_state(token, state)
        return redirect(url_for('import_grades.mapping', token=token))

    return render_template('import_grades_sheet.html', token=token, state=state)


@import_grades_bp.route('/admin/import_grades/<token>/mapping', methods=['GET', 'POST'])
@permission_required('import_grades')
def mapping(token):
    """Крок 3: зіставлення предметів і студентів (авто + вручну)."""
    state = _load_state_or_none(token)
    if not state or 'parsed' not in state:
        flash('Сесію імпорту не знайдено або вона застаріла - завантажте файл ще раз', 'warning')
        return redirect(url_for('import_grades.upload'))

    conn = get_db()
    db_subjects = [dict(r) for r in conn.execute(
        "SELECT id, name, type FROM subjects WHERE group_id = ? ORDER BY position, id", (state['group_id'],)
    ).fetchall()]
    db_students = [dict(r) for r in conn.execute(
        """SELECT id, TRIM(last_name_UA || ' ' || first_name_UA || ' ' || COALESCE(middle_name_UA,'')) AS full_name
           FROM students WHERE group_id = ? ORDER BY last_name_UA""", (state['group_id'],)
    ).fetchall()]
    group = conn.execute("SELECT name FROM groups WHERE id = ?", (state['group_id'],)).fetchone()
    conn.close()

    if not db_subjects:
        flash('В обраній групі немає предметів - спершу додайте їх (Адміністрування → Предмети)', 'warning')
        return redirect(url_for('import_grades.upload'))
    if not db_students:
        flash('В обраній групі немає студентів', 'warning')
        return redirect(url_for('import_grades.upload'))

    if request.method == 'POST':
        subj_map, stud_map = {}, {}
        for i in range(len(state['parsed']['subjects'])):
            v = request.form.get(f'subject_{i}', type=int)
            if v:
                subj_map[str(i)] = v
        for i in range(len(state['parsed']['students'])):
            v = request.form.get(f'student_{i}', type=int)
            if v:
                stud_map[str(i)] = v

        # Захист від підміни форми: лише предмети/студенти саме цієї групи
        valid_sub = {s['id'] for s in db_subjects}
        valid_stud = {s['id'] for s in db_students}
        subj_map = {k: v for k, v in subj_map.items() if v in valid_sub}
        stud_map = {k: v for k, v in stud_map.items() if v in valid_stud}

        if not subj_map or not stud_map:
            flash('Потрібно зіставити хоча б один предмет і одного студента', 'warning')
            return redirect(url_for('import_grades.mapping', token=token))

        # Захист від дублів: той самий предмет/студент не може бути
        # обраний одразу для двох різних колонок/рядків Excel (у
        # звичайному режимі це неможливо через JS на сторінці, але
        # перевіряємо і на сервері - про всяк випадок).
        if len(subj_map.values()) != len(set(subj_map.values())):
            flash('Один і той самий предмет обрано для кількох колонок Excel - виправте зіставлення', 'danger')
            return redirect(url_for('import_grades.mapping', token=token))
        if len(stud_map.values()) != len(set(stud_map.values())):
            flash('Один і той самий студент обраний для кількох рядків Excel - виправте зіставлення', 'danger')
            return redirect(url_for('import_grades.mapping', token=token))

        state['subj_map'] = subj_map
        state['stud_map'] = stud_map
        _save_state(token, state)
        return redirect(url_for('import_grades.preview', token=token))

    # GET: автоматичне зіставлення
    subj_rows = []
    for i, s in enumerate(state['parsed']['subjects']):
        mid, score, status = _match_subject(s['title'], db_subjects)
        subj_rows.append({'i': i, 'title': s['title'], 'match_id': mid, 'score': round(score, 2), 'status': status})

    stud_rows = []
    for i, s in enumerate(state['parsed']['students']):
        mid, score, status = _match_name(s['name'], db_students)
        stud_rows.append({'i': i, 'name': s['name'], 'match_id': mid, 'score': round(score, 2), 'status': status})

    return render_template(
        'import_grades_mapping.html',
        token=token, state=state, group=group,
        subj_rows=subj_rows, stud_rows=stud_rows,
        db_subjects=db_subjects, db_students=db_students,
    )


@import_grades_bp.route('/admin/import_grades/<token>/preview', methods=['GET', 'POST'])
@permission_required('import_grades')
def preview(token):
    """Крок 4: попередній перегляд; POST = підтвердження й запис у базу."""
    state = _load_state_or_none(token)
    if not state or 'subj_map' not in state:
        flash('Сесію імпорту не знайдено або вона застаріла - завантажте файл ще раз', 'warning')
        return redirect(url_for('import_grades.upload'))

    conn = get_db()
    subj_names = {r['id']: r['name'] for r in conn.execute(
        "SELECT id, name FROM subjects WHERE group_id = ?", (state['group_id'],)).fetchall()}
    stud_names = {r['id']: r['full_name'] for r in conn.execute(
        """SELECT id, TRIM(last_name_UA || ' ' || first_name_UA || ' ' || COALESCE(middle_name_UA,'')) AS full_name
           FROM students WHERE group_id = ?""", (state['group_id'],)).fetchall()}

    parsed = state['parsed']
    subj_map = {int(k): v for k, v in state['subj_map'].items()}
    stud_map = {int(k): v for k, v in state['stud_map'].items()}

    # Зібрати план імпорту
    plan = []           # (student_id, subject_id, value)
    for key, val in parsed['grades'].items():
        si_str, ci_str = key.split(',')
        stud_i, subj_i = int(si_str), int(ci_str)
        if stud_i in stud_map and subj_i in subj_map:
            plan.append((stud_map[stud_i], subj_map[subj_i], val))

    existing = {}
    for row in conn.execute(
            "SELECT student_id, subject_id, grade FROM grades WHERE student_id IN (%s)" %
            ','.join('?' * len(set(stud_map.values()))), list(set(stud_map.values()))).fetchall():
        existing[(row['student_id'], row['subject_id'])] = row['grade']

    overwrites = sum(1 for sid, subid, _ in plan if (sid, subid) in existing)

    if request.method == 'POST':
        inserted = updated = 0
        try:
            for sid, subid, val in plan:
                if (sid, subid) in existing:
                    conn.execute("UPDATE grades SET grade = ? WHERE student_id = ? AND subject_id = ?",
                                 (str(val), sid, subid))
                    updated += 1
                else:
                    conn.execute("INSERT INTO grades (student_id, subject_id, grade) VALUES (?, ?, ?)",
                                 (sid, subid, str(val)))
                    inserted += 1
            conn.commit()
            log_action(
                current_username(),
                f"імпортував оцінки з Excel ({state.get('filename','')}, аркуш «{state.get('sheet','')}»)",
                group_ids=[state['group_id']],
                details=f"додано {inserted}, оновлено {updated}, студентів {len(set(stud_map.values()))}, предметів {len(set(subj_map.values()))}"
            )
            flash(f"Імпорт завершено: додано {inserted} оцінок, оновлено {updated}", 'success')
        except Exception as e:
            conn.rollback()
            logger.error(f"Імпорт оцінок: помилка запису: {e}", exc_info=True)
            flash(f'Помилка при записі оцінок: {e}', 'danger')
        finally:
            conn.close()
            # Прибираємо тимчасові файли сесії
            for p in (_state_path(token), _xlsx_path(token)):
                try:
                    os.remove(p)
                except OSError:
                    pass
        return redirect(url_for('students.student_list'))

    conn.close()

    # Побудова матриці для перегляду
    mapped_subjects = [(i, subj_map[i], subj_names.get(subj_map[i], '?')) for i in sorted(subj_map)]
    rows = []
    for stud_i in sorted(stud_map):
        cells = []
        for subj_i, subid, _ in mapped_subjects:
            key = f"{stud_i},{subj_i}"
            val = parsed['grades'].get(key)
            over = val is not None and (stud_map[stud_i], subid) in existing
            cells.append({'val': val, 'overwrite': over})
        rows.append({
            'excel_name': parsed['students'][stud_i]['name'],
            'db_name': stud_names.get(stud_map[stud_i], '?'),
            'cells': cells,
        })

    skipped_subjects = [s['title'] for i, s in enumerate(parsed['subjects']) if i not in subj_map]
    skipped_students = [s['name'] for i, s in enumerate(parsed['students']) if i not in stud_map]

    return render_template(
        'import_grades_preview.html',
        token=token, state=state,
        subjects=[name for _, _, name in mapped_subjects],
        rows=rows, total=len(plan), overwrites=overwrites,
        skipped_subjects=skipped_subjects, skipped_students=skipped_students,
    )