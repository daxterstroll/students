from flask import Blueprint, render_template, request, redirect, url_for, session, flash, send_file
from datetime import datetime
from zipfile import ZipFile
import os
import sqlite3
from werkzeug.security import generate_password_hash
from db import get_db
from utils import log_action, permission_required
from gen_docx import gen_doc
import logging
import openpyxl
from werkzeug.utils import secure_filename
import pandas as pd
import json
import locale
import uuid
import re
from openpyxl import load_workbook
from rapidfuzz import process, fuzz
from deep_translator import GoogleTranslator
from openpyxl.utils.exceptions import InvalidFileException
import time


from utils import get_available_templates

translator = GoogleTranslator(source="auto", target="en")
translation_cache = {}

admin_bp = Blueprint('admin', __name__)

PERMISSIONS = [
    'manage_users',
    'view_logs',
    'group_export',
    'import_from_excel',
    'manage_education_documents',
    'study_periods',
    'manage_groups',
    'manage_subjects',
    'manage_activities',
    'import_subjects',
    'archive',
    'manage_students',
    'manage_accreditations',
    'manage_diplomas',
    'import_education_docs'
]






@admin_bp.route('/admin/study_periods/bulk_assign', methods=['GET', 'POST'])
@permission_required('study_periods')
def manage_study_periods_bulk_assign():
    db = get_db()
    cursor = db.cursor()

    # -------------------- POST --------------------
    if request.method == 'POST':
        mode = request.form.get('mode')
        group_id = request.form.get('group_id', type=int)
        student_ids_raw = request.form.get('student_ids_hidden', '')
        extra_student_ids = [int(x) for x in student_ids_raw.split(',') if x.strip().isdigit()]

        target_student_ids = set(extra_student_ids)

        if group_id and mode in ('whole_group', 'some_from_group'):
            cursor.execute("""
                SELECT id FROM students WHERE group_id = ? AND archived = FALSE
            """, (group_id,))
            target_student_ids.update(r['id'] for r in cursor.fetchall())

        if not target_student_ids:
            flash('Не обрано жодного студента', 'danger')
            return redirect(url_for('admin.manage_study_periods_bulk_assign'))

        filiya = (request.form.get('filiya') or '').strip()
        filiya_en = (request.form.get('filiya_en') or '').strip() or None
        group_name = (request.form.get('group_name') or '').strip() or None
        start_date = (request.form.get('start_date') or '').strip() or None
        end_date = (request.form.get('end_date') or '').strip() or None
        period_order = request.form.get('period_order', type=int) or 0
        note = (request.form.get('note') or '').strip() or None

        if not filiya:
            flash('Потрібно вказати філію', 'danger')
            return redirect(url_for('admin.manage_study_periods_bulk_assign',
                                  group_id=group_id or '',
                                  student_ids=','.join(map(str, extra_student_ids)),
                                  mode=mode))

        added = 0
        try:
            for student_id in target_student_ids:
                cursor.execute("""
                    INSERT INTO student_study_periods
                        (student_id, filiya, filiya_en, group_name, start_date, end_date, period_order, note)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """, (student_id, filiya, filiya_en, group_name, start_date, end_date, period_order, note))
                added += 1

            db.commit()
            log_action(
                session.get('username', 'невідомо'),
                f"масово присвоїв період навчання ({filiya}) для {added} студент(ів)"
            )
            flash(f'Період навчання успішно додано {added} студент(ам)', 'success')
        except sqlite3.Error as e:
            db.rollback()
            flash(f'Помилка збереження: {e}', 'danger')

        return redirect(url_for('admin.manage_study_periods_bulk_assign',
                              group_id=group_id or '',
                              student_ids=','.join(map(str, extra_student_ids)),
                              mode=mode))

    # -------------------- GET --------------------
    cursor.execute("SELECT id, name FROM groups WHERE archived = FALSE ORDER BY name")
    groups = cursor.fetchall()

    # Всі студенти з інформацією про групу (для JS-фільтрації)
    cursor.execute("""
        SELECT s.id, s.last_name_UA, s.first_name_UA, s.middle_name_UA, 
               s.group_id, g.name AS group_name
        FROM students s
        LEFT JOIN groups g ON s.group_id = g.id
        WHERE s.archived = FALSE
        ORDER BY s.last_name_UA, s.first_name_UA
    """)
    all_students = cursor.fetchall()

    # Параметри
    mode = request.args.get('mode', 'any_students')
    group_id = request.args.get('group_id', type=int)
    student_ids_param = request.args.get('student_ids', '')
    extra_student_ids = [int(x) for x in student_ids_param.split(',') if x.strip().isdigit()]

    target_student_ids = set(extra_student_ids)

    if group_id and mode in ('whole_group', 'some_from_group'):
        cursor.execute("""
            SELECT id FROM students WHERE group_id = ? AND archived = FALSE
        """, (group_id,))
        target_student_ids.update(r['id'] for r in cursor.fetchall())

    # Прев'ю
    target_students = []
    if target_student_ids:
        placeholders = ','.join('?' for _ in target_student_ids)
        cursor.execute(f"""
            SELECT s.id, s.last_name_UA, s.first_name_UA, s.middle_name_UA, g.name AS group_name
            FROM students s
            LEFT JOIN groups g ON s.group_id = g.id
            WHERE s.id IN ({placeholders})
            ORDER BY s.last_name_UA, s.first_name_UA
        """, tuple(target_student_ids))
        target_students_rows = cursor.fetchall()

        cursor.execute(f"""
            SELECT student_id, filiya, filiya_en, group_name, start_date, end_date, period_order, note
            FROM student_study_periods
            WHERE student_id IN ({placeholders})
            ORDER BY student_id, period_order ASC, start_date ASC
        """, tuple(target_student_ids))
        periods_rows = cursor.fetchall()

        periods_by_student = {}
        for p in periods_rows:
            periods_by_student.setdefault(p['student_id'], []).append(p)

        for s in target_students_rows:
            s_dict = dict(s)
            s_dict['existing_periods'] = periods_by_student.get(s['id'], [])
            target_students.append(s_dict)

    return render_template(
        'manage_study_periods_bulk_assign.html',
        groups=groups,
        all_students=all_students,
        selected_group_id=group_id,
        selected_student_ids=extra_student_ids,
        target_students=target_students,
        selected_mode=mode
    )


@admin_bp.route('/admin/manage_diplomas', methods=['GET', 'POST'])
@permission_required('manage_diplomas')
def manage_diplomas():
    conn = get_db()
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    if request.method == "POST":
        group_id = request.form.get("group_id")
        cursor.execute("SELECT s.id FROM students s WHERE s.group_id = ?", (group_id,))
        students = cursor.fetchall()

        for student in students:
            student_id = student['id']
            diploma_number = request.form.get(f'diploma_number_{student_id}', '').strip()
            appendix_number = request.form.get(f'appendix_number_{student_id}', '').strip()
            if diploma_number:
                diploma_number = diploma_number.zfill(6)
            cursor.execute("SELECT id FROM diplomas WHERE student_id=?", (student_id,))
            exists = cursor.fetchone()
            if exists:
                cursor.execute("""
                    UPDATE diplomas SET diploma_number=?, appendix_number=? WHERE student_id=?
                """, (diploma_number, appendix_number, student_id))
            else:
                cursor.execute("""
                    INSERT INTO diplomas(student_id, diploma_number, appendix_number) VALUES (?, ?, ?)
                """, (student_id, diploma_number, appendix_number))

        conn.commit()

        group_row = cursor.execute("SELECT name FROM groups WHERE id=?", (group_id,)).fetchone()
        group_name = group_row['name'] if group_row else f"ID {group_id}"
        log_action(
            session.get('username', 'невідомо'),
            f"зберіг дипломи для групи: {group_name} (ID {group_id})",
            details=f"студентів оброблено: {len(students)}"
        )

        flash("Дані збережено")
        return redirect(url_for('admin.manage_diplomas', group_id=group_id))

    cursor.execute("SELECT id, name, start_year FROM groups WHERE archived = FALSE ORDER BY name")
    groups = cursor.fetchall()
    selected_group = request.args.get("group_id")
    if selected_group is not None:
        selected_group = int(selected_group)

    students = []
    if selected_group:
        cursor.execute("""
            SELECT s.id, s.last_name_UA, s.first_name_UA, s.middle_name_UA,
                   d.diploma_number, d.appendix_number
            FROM students s
            LEFT JOIN diplomas d ON s.id = d.student_id
            WHERE s.group_id = ?
        """, (selected_group,))
        students = cursor.fetchall()
        try:
            locale.setlocale(locale.LC_COLLATE, 'uk_UA.UTF-8')
        except locale.Error:
            try:
                locale.setlocale(locale.LC_COLLATE, 'Ukrainian_Ukraine.1251')
            except locale.Error:
                pass
        students = sorted(
            students,
            key=lambda s: locale.strxfrm(f"{s['last_name_UA']} {s['first_name_UA']} {s['middle_name_UA']}")
        )

    return render_template("manage_diplomas.html", groups=groups, students=students, selected_group=selected_group)


@admin_bp.route('/admin/manage_accreditations', methods=['GET', 'POST'])
@permission_required('manage_accreditations')
def manage_accreditations():
    conn = get_db()
    cursor = conn.cursor()

    if request.method == 'POST' and 'add' in request.form:
        degree = request.form.get('degree')
        specialty = request.form.get('specialty')
        text_ua = request.form.get('text_ua')
        text_en = request.form.get('text_en')
        cursor.execute("""
            INSERT INTO accreditations (degree, specialty, text_ua, text_en) VALUES (?, ?, ?, ?)
        """, (degree, specialty, text_ua, text_en))
        conn.commit()
        log_action(
            session.get('username', 'невідомо'),
            f"додав акредитацію: {degree} / {specialty}"
        )
        return redirect(url_for('admin.manage_accreditations'))

    if request.method == 'POST' and 'edit' in request.form:
        acc_id = request.form.get('id')
        degree = request.form.get('degree')
        specialty = request.form.get('specialty')
        text_ua = request.form.get('text_ua')
        text_en = request.form.get('text_en')
        cursor.execute("""
            UPDATE accreditations SET degree=?, specialty=?, text_ua=?, text_en=? WHERE id=?
        """, (degree, specialty, text_ua, text_en, acc_id))
        conn.commit()
        log_action(
            session.get('username', 'невідомо'),
            f"редагував акредитацію ID {acc_id}: {degree} / {specialty}"
        )
        return redirect(url_for('admin.manage_accreditations'))

    if request.method == 'POST' and 'delete' in request.form:
        acc_id = request.form.get('id')
        row = cursor.execute("SELECT degree, specialty FROM accreditations WHERE id=?", (acc_id,)).fetchone()
        cursor.execute("DELETE FROM accreditations WHERE id=?", (acc_id,))
        conn.commit()
        log_action(
            session.get('username', 'невідомо'),
            f"ВИДАЛИВ акредитацію ID {acc_id}: {row[0] if row else ''} / {row[1] if row else ''}"
        )
        return redirect(url_for('admin.manage_accreditations'))

    cursor.execute("SELECT id, degree, specialty, text_ua, text_en FROM accreditations ORDER BY degree, specialty")
    accreditations = cursor.fetchall()
    cursor.execute("SELECT DISTINCT specialty FROM groups WHERE archived = FALSE ORDER BY specialty")
    groups = cursor.fetchall()

    return render_template('manage_accreditations.html', accreditations=accreditations, groups=groups)


@admin_bp.route('/admin/manage_education_documents', methods=['GET', 'POST'])
@permission_required('manage_education_documents')
def manage_education_documents():
    db = get_db()
    cursor = db.cursor()

    cursor.execute("""
        SELECT id, last_name_UA, first_name_UA FROM students
        WHERE archived = FALSE ORDER BY last_name_UA, first_name_UA
    """)
    students = cursor.fetchall()

    cursor.execute("SELECT id, name FROM groups WHERE archived = FALSE ORDER BY name")
    groups = cursor.fetchall()

    # --- Отримуємо студента, якщо перейшли з картки ---
    student = None
    student_id_param = request.args.get('student_id', type=int)
    if student_id_param:
        cursor.execute("""
            SELECT id, last_name_UA, first_name_UA, middle_name_UA, group_id
            FROM students WHERE id = ?
        """, (student_id_param,))
        srow = cursor.fetchone()
        if srow:
            student = dict(srow)
            cursor.execute("SELECT name FROM groups WHERE id = ?", (student['group_id'],))
            grow = cursor.fetchone()
            student['group_name'] = grow['name'] if grow else ''

    selected_group_id = request.args.get('group_id', type=int)
    if not selected_group_id and student:
        selected_group_id = student['group_id']

    students_without_docs = []
    if selected_group_id:
        cursor.execute("""
            SELECT s.id, s.last_name_UA, s.first_name_UA
            FROM students s
            WHERE s.group_id = ? AND s.archived = FALSE
              AND s.id NOT IN (SELECT student_id FROM education_documents)
            ORDER BY s.last_name_UA, s.first_name_UA
        """, (selected_group_id,))
        students_without_docs = cursor.fetchall()

    # Основний запит документів
    cursor.execute("""
        SELECT g.id AS group_id, g.name AS group_name, s.id AS student_id,
               s.last_name_UA, s.first_name_UA,
               ed.id AS doc_id, ed.document_type, ed.document_type_en, ed.document_number,
               ed.institution_name, ed.institution_name_en, ed.country, ed.country_en, ed.completion_date,
               fed.reference_number, fed.reference_institution, fed.reference_institution_en,
               fed.reference_country, fed.reference_country_en, fed.reference_issue_date,
               fed.recognition_certificate_number, fed.recognition_issuer,
               fed.recognition_issuer_en, fed.recognition_date
        FROM education_documents ed
        INNER JOIN students s ON ed.student_id = s.id
        INNER JOIN groups g ON s.group_id = g.id
        LEFT JOIN foreign_education_docs fed ON ed.id = fed.education_doc_id
        WHERE s.archived = FALSE AND g.archived = FALSE
        ORDER BY g.name, s.last_name_UA, s.first_name_UA, ed.id
    """)
    rows = cursor.fetchall()

    documents_by_group = {}
    for row in rows:
        gid = row['group_id']
        if gid not in documents_by_group:
            documents_by_group[gid] = {'group_name': row['group_name'], 'docs': []}
        documents_by_group[gid]['docs'].append(row)

    sorted_documents_by_group = sorted(documents_by_group.items(), key=lambda x: x[1]['group_name'])

    # Перевіряємо doc_id для студента
    student_doc_id = None
    if student:
        for gid, gdata in documents_by_group.items():
            for doc in gdata['docs']:
                if doc['student_id'] == student['id']:
                    student_doc_id = doc['doc_id']
                    break
            if student_doc_id:
                break

    # ====================== POST ======================
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'delete':
            doc_id = request.form.get('doc_id')
            try:
                cursor.execute("DELETE FROM foreign_education_docs WHERE education_doc_id = ?", (doc_id,))
                cursor.execute("DELETE FROM education_documents WHERE id = ?", (doc_id,))
                db.commit()
                log_action(session.get('username', 'невідомо'), f"ВИДАЛИВ документ про освіту ID {doc_id}")
                flash('Документ успішно видалено', 'success')
            except sqlite3.Error as e:
                db.rollback()
                flash(f'Помилка видалення: {e}', 'danger')

        elif action == 'edit':
            doc_id = request.form.get('doc_id')
            student_id = request.form.get('student_id')

            country = request.form.get('country') or ''
            country_en = request.form.get('country_en') or ''

            try:
                cursor.execute("SELECT student_id FROM education_documents WHERE id = ?", (doc_id,))
                existing = cursor.fetchone()
                if not existing:
                    flash('Документ не знайдено', 'danger')
                    return redirect(url_for('admin.manage_education_documents', group_id=selected_group_id))

                if student_id:
                    cursor.execute("SELECT id FROM students WHERE id = ? AND archived = FALSE", (student_id,))
                    if not cursor.fetchone():
                        flash('Обраний студент не існує або заархівований', 'danger')
                        return redirect(url_for('admin.manage_education_documents', group_id=selected_group_id))
                else:
                    student_id = existing[0]

                # Оновлення основного документа (зберігаємо country як ввели)
                cursor.execute("""
                    UPDATE education_documents SET
                        student_id=?, document_type=?, document_type_en=?, document_number=?,
                        institution_name=?, institution_name_en=?, country=?, country_en=?, completion_date=?
                    WHERE id=?
                """, (student_id, request.form.get('document_type'), request.form.get('document_type_en'),
                      request.form.get('document_number'), request.form.get('institution_name'),
                      request.form.get('institution_name_en'), country, country_en,
                      request.form.get('completion_date'), doc_id))

                # Foreign fields
                reference_number = request.form.get('reference_number') or None
                reference_institution = request.form.get('reference_institution') or None
                reference_institution_en = request.form.get('reference_institution_en') or None
                reference_country = request.form.get('reference_country') or None
                reference_country_en = request.form.get('reference_country_en') or None
                reference_issue_date = request.form.get('reference_issue_date') or None
                recognition_certificate_number = request.form.get('recognition_certificate_number') or None
                recognition_issuer = request.form.get('recognition_issuer') or None
                recognition_issuer_en = request.form.get('recognition_issuer_en') or None
                recognition_date = request.form.get('recognition_date') or None

                # Рішення приймаємо тільки на основі того, чи заповнені самі поля довідки/визнання,
                # а НЕ на основі значення country. Це прибирає випадкове видалення довідки
                # через неправильно передане/незмінене поле "Країна".
                has_foreign_data = any([
                    reference_number, reference_institution, reference_country,
                    reference_issue_date, recognition_certificate_number,
                    recognition_issuer, recognition_date
                ])

                cursor.execute("SELECT id FROM foreign_education_docs WHERE education_doc_id = ?", (doc_id,))
                foreign_exists = cursor.fetchone()

                if foreign_exists:
                    if has_foreign_data:
                        cursor.execute("""
                            UPDATE foreign_education_docs SET
                                reference_number=?, reference_institution=?, reference_institution_en=?,
                                reference_country=?, reference_country_en=?, reference_issue_date=?,
                                recognition_certificate_number=?, recognition_issuer=?,
                                recognition_issuer_en=?, recognition_date=?
                            WHERE education_doc_id=?
                        """, (reference_number, reference_institution, reference_institution_en,
                              reference_country, reference_country_en, reference_issue_date,
                              recognition_certificate_number, recognition_issuer,
                              recognition_issuer_en, recognition_date, doc_id))
                    else:
                        # Видаляємо запис тільки якщо користувач сам очистив усі поля довідки/визнання
                        cursor.execute("DELETE FROM foreign_education_docs WHERE education_doc_id = ?", (doc_id,))
                else:
                    if has_foreign_data:
                        cursor.execute("""
                            INSERT INTO foreign_education_docs (
                                education_doc_id, reference_number, reference_institution, reference_institution_en,
                                reference_country, reference_country_en, reference_issue_date,
                                recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (doc_id, reference_number, reference_institution, reference_institution_en,
                              reference_country, reference_country_en, reference_issue_date,
                              recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date))

                db.commit()
                log_action(
                    session.get('username', 'невідомо'),
                    f"редагував документ про освіту ID {doc_id}",
                    details=f"країна: {country}"
                )
                flash('Документ успішно оновлено', 'success')

            except sqlite3.Error as e:
                db.rollback()
                flash(f'Помилка при редагуванні: {e}', 'danger')
            except Exception as e:
                db.rollback()
                flash(f'Непередбачена помилка: {str(e)}', 'danger')

            return redirect(url_for('admin.manage_education_documents', group_id=selected_group_id))

        # === Додавання нового документа ===
        else:
            country = request.form.get('country') or ''
            country_en = request.form.get('country_en') or ''

            student_id = request.form.get('student_id')
            document_type = request.form.get('document_type')
            document_type_en = request.form.get('document_type_en')
            document_number = request.form.get('document_number')
            institution_name = request.form.get('institution_name')
            institution_name_en = request.form.get('institution_name_en')
            completion_date = request.form.get('completion_date')

            reference_number = request.form.get('reference_number') or None
            reference_institution = request.form.get('reference_institution') or None
            reference_institution_en = request.form.get('reference_institution_en') or None
            reference_country = request.form.get('reference_country') or None
            reference_country_en = request.form.get('reference_country_en') or None
            reference_issue_date = request.form.get('reference_issue_date') or None
            recognition_certificate_number = request.form.get('recognition_certificate_number') or None
            recognition_issuer = request.form.get('recognition_issuer') or None
            recognition_issuer_en = request.form.get('recognition_issuer_en') or None
            recognition_date = request.form.get('recognition_date') or None

            try:
                if not student_id:
                    raise ValueError("Не обрано студента")

                cursor.execute("""
                    INSERT INTO education_documents (
                        student_id, document_type, document_type_en, document_number,
                        institution_name, institution_name_en, country, country_en, completion_date
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (student_id, document_type, document_type_en, document_number,
                      institution_name, institution_name_en, country, country_en, completion_date))

                education_doc_id = cursor.lastrowid

                has_foreign_data = any([
                    reference_number, reference_institution, reference_country,
                    reference_issue_date, recognition_certificate_number,
                    recognition_issuer, recognition_date
                ])

                if has_foreign_data:
                    cursor.execute("""
                        INSERT INTO foreign_education_docs (
                            education_doc_id, reference_number, reference_institution, reference_institution_en,
                            reference_country, reference_country_en, reference_issue_date,
                            recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (education_doc_id, reference_number, reference_institution, reference_institution_en,
                          reference_country, reference_country_en, reference_issue_date,
                          recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date))

                db.commit()
                log_action(
                    session.get('username', 'невідомо'),
                    f"додав документ про освіту: {document_type} №{document_number}",
                    details=f"країна: {country}"
                )
                flash('Документ успішно додано', 'success')

            except (sqlite3.Error, ValueError) as e:
                db.rollback()
                flash(f'Помилка при додаванні: {str(e)}', 'danger')

        return redirect(url_for('admin.manage_education_documents', group_id=selected_group_id))

    cursor.close()
    return render_template(
        'manage_education_documents.html',
        groups=groups,
        selected_group_id=selected_group_id,
        students_without_docs=students_without_docs,
        documents_by_group=sorted_documents_by_group,
        students=students,
        student=student,
        student_doc_id=student_doc_id
    )
    
@admin_bp.route('/admin/manage_groups', methods=['GET', 'POST'])
@permission_required('manage_groups')
def manage_groups():
    conn = get_db()
    conn.row_factory = sqlite3.Row

    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'add':
            name = request.form.get('name')
            start_year = request.form.get('start_year')
            study_form = request.form.get('study_form')
            program_credits = request.form.get('program_credits')
            qualification_name = request.form.get('qualification_name')
            degree_level = request.form.get('degree_level')
            specialty = request.form.get('specialty')
            educational_program = request.form.get('educational_program')
            knowledge_area = request.form.get('knowledge_area')
            qualification_name_en = request.form.get('qualification_name_en')
            degree_level_en = request.form.get('degree_level_en')
            specialty_en = request.form.get('specialty_en')
            educational_program_en = request.form.get('educational_program_en')
            knowledge_area_en = request.form.get('knowledge_area_en')
            institution_name_and_status = request.form.get('institution_name_and_status')
            institution_name_and_status_en = request.form.get('institution_name_and_status_en')
            entry_requirements = request.form.get('entry_requirements')
            entry_requirements_en = request.form.get('entry_requirements_en')
            learning_outcomes = request.form.get('learning_outcomes')
            learning_outcomes_en = request.form.get('learning_outcomes_en')
            program_includes = request.form.get('program_includes')
            program_includes_en = request.form.get('program_includes_en')

            required_fields = [name, start_year, study_form, program_credits]
            if not all(required_fields):
                flash("Усі поля мають бути заповнені.", "error")
            elif study_form not in ['Денна', 'Заочна']:
                flash("Форма навчання має бути 'Денна' або 'Заочна'.", "error")
            elif program_credits not in ['90', '120', '180', '240']:
                flash("Кількість кредитів має бути 90, 120, 180 або 240.", "error")
            else:
                try:
                    start_year = int(start_year)
                    program_credits = int(program_credits)
                    current_year = datetime.now().year
                    if start_year < 2000 or start_year > current_year:
                        flash(f"Рік початку навчання має бути між 2000 і {current_year}.", "error")
                    else:
                        conn.execute("""
                            INSERT INTO groups (
                                name, start_year, study_form, program_credits,
                                qualification_name, degree_level, specialty, educational_program, knowledge_area,
                                qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                                institution_name_and_status, institution_name_and_status_en,
                                entry_requirements, entry_requirements_en,
                                learning_outcomes, learning_outcomes_en, program_includes, program_includes_en
                            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """, (name, start_year, study_form, program_credits,
                              qualification_name, degree_level, specialty, educational_program, knowledge_area,
                              qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                              institution_name_and_status, institution_name_and_status_en,
                              entry_requirements, entry_requirements_en,
                              learning_outcomes, learning_outcomes_en, program_includes, program_includes_en))
                        conn.commit()
                        flash("Групу додано успішно.", "success")
                        log_action(
                            session.get('username', 'невідомо'),
                            f"додав групу: {name} ({start_year}, {study_form}, {program_credits} кредитів)",
                            details=f"спеціальність: {specialty or 'не вказано'}, ступінь: {degree_level or 'не вказано'}"
                        )
                except ValueError:
                    flash("Рік початку навчання або кредити мають бути числами.", "error")
                except sqlite3.IntegrityError:
                    flash("Група з таким назвою та роком початку навчання вже існує.", "error")

        elif action == 'edit':
            group_id = request.form.get('group_id')
            name = request.form.get('name')
            start_year = request.form.get('start_year')
            study_form = request.form.get('study_form')
            program_credits = request.form.get('program_credits')
            qualification_name = request.form.get('qualification_name')
            degree_level = request.form.get('degree_level')
            specialty = request.form.get('specialty')
            educational_program = request.form.get('educational_program')
            knowledge_area = request.form.get('knowledge_area')
            qualification_name_en = request.form.get('qualification_name_en')
            degree_level_en = request.form.get('degree_level_en')
            specialty_en = request.form.get('specialty_en')
            educational_program_en = request.form.get('educational_program_en')
            knowledge_area_en = request.form.get('knowledge_area_en')
            institution_name_and_status = request.form.get('institution_name_and_status')
            institution_name_and_status_en = request.form.get('institution_name_and_status_en')
            entry_requirements = request.form.get('entry_requirements')
            entry_requirements_en = request.form.get('entry_requirements_en')
            learning_outcomes = request.form.get('learning_outcomes')
            learning_outcomes_en = request.form.get('learning_outcomes_en')
            program_includes = request.form.get('program_includes')
            program_includes_en = request.form.get('program_includes_en')

            required_fields = [group_id, name, start_year, study_form, program_credits]
            if not all(required_fields):
                flash("Усі поля мають бути заповнені.", "error")
            elif study_form not in ['Денна', 'Заочна']:
                flash("Форма навчання має бути 'Денна' або 'Заочна'.", "error")
            elif program_credits not in ['90', '120', '180', '240']:
                flash("Кількість кредитів має бути 90, 120, 180 або 240.", "error")
            else:
                try:
                    start_year = int(start_year)
                    program_credits = int(program_credits)
                    current_year = datetime.now().year
                    if start_year < 2000 or start_year > current_year:
                        flash(f"Рік початку навчання має бути між 2000 і {current_year}.", "error")
                    else:
                        conn.execute("""
                            UPDATE groups SET
                                name=?, start_year=?, study_form=?, program_credits=?,
                                qualification_name=?, degree_level=?, specialty=?,
                                educational_program=?, knowledge_area=?,
                                qualification_name_en=?, degree_level_en=?, specialty_en=?,
                                educational_program_en=?, knowledge_area_en=?,
                                institution_name_and_status=?, institution_name_and_status_en=?,
                                entry_requirements=?, entry_requirements_en=?,
                                learning_outcomes=?, learning_outcomes_en=?,
                                program_includes=?, program_includes_en=?
                            WHERE id=?
                        """, (name, start_year, study_form, program_credits,
                              qualification_name, degree_level, specialty, educational_program, knowledge_area,
                              qualification_name_en, degree_level_en, specialty_en, educational_program_en, knowledge_area_en,
                              institution_name_and_status, institution_name_and_status_en,
                              entry_requirements, entry_requirements_en,
                              learning_outcomes, learning_outcomes_en, program_includes, program_includes_en,
                              group_id))
                        conn.commit()
                        flash("Групу відредаговано успішно.", "success")
                        log_action(
                            session.get('username', 'невідомо'),
                            f"редагував групу: {name} (ID {group_id})",
                            details=f"рік: {start_year}, форма: {study_form}, кредити: {program_credits}"
                        )
                except ValueError:
                    flash("Рік початку навчання або кредити мають бути числами.", "error")
                except sqlite3.IntegrityError:
                    flash("Група з таким назвою та роком початку навчання вже існує.", "error")

        elif action == 'delete':
            group_id = request.form.get('group_id')
            group_row = conn.execute("SELECT name, start_year FROM groups WHERE id=?", (group_id,)).fetchone()
            related_data = conn.execute("""
                SELECT (SELECT COUNT(*) FROM students WHERE group_id=?) +
                       (SELECT COUNT(*) FROM subjects WHERE group_id=?) +
                       (SELECT COUNT(*) FROM practices WHERE group_id=?) +
                       (SELECT COUNT(*) FROM courseworks WHERE group_id=?) +
                       (SELECT COUNT(*) FROM attestations WHERE group_id=?) AS total
            """, (group_id, group_id, group_id, group_id, group_id)).fetchone()['total']

            if related_data > 0:
                flash("Неможливо видалити групу, оскільки вона має пов'язані дані.", "error")
            else:
                conn.execute("DELETE FROM groups WHERE id=?", (group_id,))
                conn.commit()
                flash("Групу видалено успішно.", "success")
                log_action(
                    session.get('username', 'невідомо'),
                    f"ВИДАЛИВ групу: {group_row['name']} ({group_row['start_year']}) (ID {group_id})"
                )

    groups = conn.execute("""
        SELECT g.id, g.name, g.start_year, g.study_form, g.program_credits,
               g.qualification_name, g.degree_level, g.specialty, g.educational_program, g.knowledge_area,
               g.qualification_name_en, g.degree_level_en, g.specialty_en, g.educational_program_en, g.knowledge_area_en,
               g.institution_name_and_status, g.institution_name_and_status_en,
               g.entry_requirements, g.entry_requirements_en,
               g.learning_outcomes, g.learning_outcomes_en, g.program_includes, g.program_includes_en,
               g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS display_name,
               (SELECT COUNT(*) FROM students s WHERE s.group_id = g.id) AS student_count
        FROM groups g WHERE g.archived = FALSE ORDER BY g.id, g.start_year
    """).fetchall()

    conn.close()
    return render_template("manage_groups.html", groups=groups)


@admin_bp.route('/admin/manage_subjects', methods=['GET', 'POST'])
@permission_required('manage_subjects')
def manage_subjects():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups WHERE archived = FALSE ORDER BY name, start_year
    """)
    groups = cursor.fetchall()

    selected_group_id = request.args.get('group_id')
    subjects = []
    students = []
    grades = []
    selected_subject_id = request.args.get('subject_id')

    if selected_group_id:
        cursor.execute('SELECT * FROM subjects WHERE group_id = ? ORDER BY position', (selected_group_id,))
        subjects = cursor.fetchall()
        if selected_subject_id:
            cursor.execute('SELECT * FROM students WHERE group_id = ?', (selected_group_id,))
            students = cursor.fetchall()
            import locale
            locale.setlocale(locale.LC_COLLATE, 'uk_UA.UTF-8')
            students = sorted(students, key=lambda s: locale.strxfrm(
                f"{s['last_name_UA']} {s['first_name_UA']} {s['middle_name_UA']}"))
            cursor.execute('SELECT id, student_id, subject_id, grade FROM grades WHERE subject_id = ?', (selected_subject_id,))
            grades = cursor.fetchall()

    if request.method == 'POST':
        action = request.form['action']
        group_id = request.form['group_id']

        if action == 'add':
            try:
                code = request.form['code'].strip()
                name = request.form['name'].strip()
                credits = int(request.form['credits'])
                type_ = request.form['type']
                position = int(request.form['position'])
                if not code or not name or credits < 1 or position < 1 or type_ not in ['Залік', 'Екзамен']:
                    flash('Некорректные данные предмета', 'error')
                else:
                    cursor.execute('SELECT MAX(position) FROM subjects WHERE group_id = ?', (group_id,))
                    max_position = cursor.fetchone()[0] or 0
                    if position <= max_position:
                        cursor.execute('UPDATE subjects SET position = position + 1 WHERE position >= ? AND group_id = ?', (position, group_id))
                    cursor.execute('INSERT INTO subjects (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)',
                                   (code, name, credits, type_, position, group_id))
                    conn.commit()
                    log_action(session.get('username', 'невідомо'),
                               f"додав предмет: {code} — {name} (група ID {group_id})",
                               details=f"кредити: {credits}, тип: {type_}, позиція: {position}")
                    flash(f'Добавлен предмет {code}', 'success')
            except (KeyError, ValueError):
                flash('Некорректные данные предмета', 'error')

        elif action == 'edit':
            try:
                subject_id = request.form['subject_id']
                code = request.form['code'].strip()
                name = request.form['name'].strip()
                credits = int(request.form['credits'])
                type_ = request.form['type']
                position = int(request.form['position'])
                if not code or not name or credits < 1 or position < 1 or type_ not in ['Залік', 'Екзамен']:
                    flash('Некорректные данные предмета', 'error')
                else:
                    cursor.execute('SELECT position FROM subjects WHERE id = ? AND group_id = ?', (subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position = 0 WHERE id = ? AND group_id = ?', (subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position = position + 1 WHERE position >= ? AND group_id = ? AND id != ?', (position, group_id, subject_id))
                    cursor.execute('UPDATE subjects SET code=?, name=?, credits=?, type=?, position=? WHERE id=? AND group_id=?',
                                   (code, name, credits, type_, position, subject_id, group_id))
                    cursor.execute('SELECT id, position FROM subjects WHERE group_id=? ORDER BY position, id', (group_id,))
                    for i, subj in enumerate(cursor.fetchall(), 1):
                        if subj['position'] != i:
                            cursor.execute('UPDATE subjects SET position=? WHERE id=? AND group_id=?', (i, subj['id'], group_id))
                    conn.commit()
                    log_action(session.get('username', 'невідомо'),
                               f"редагував предмет: {code} — {name} (група ID {group_id})",
                               details=f"кредити: {credits}, тип: {type_}, позиція: {position}")
                    flash(f'Обновлен предмет {code}', 'success')
            except (KeyError, ValueError):
                flash('Некорректные данные предмета', 'error')

        elif action == 'delete':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT COUNT(*) FROM grades WHERE subject_id=?', (subject_id,))
                grade_count = cursor.fetchone()[0]
                if grade_count > 0:
                    cursor.execute('DELETE FROM grades WHERE subject_id=?', (subject_id,))
                cursor.execute('SELECT position, code, name FROM subjects WHERE id=? AND group_id=?', (subject_id, group_id))
                position_row = cursor.fetchone()
                if position_row is None:
                    flash('Предмет не найден', 'error')
                else:
                    position = position_row[0]
                    subj_code = position_row[1]
                    subj_name = position_row[2]
                    cursor.execute('DELETE FROM subjects WHERE id=? AND group_id=?', (subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position=position-1 WHERE position>? AND group_id=?', (position, group_id))
                    conn.commit()
                    log_action(session.get('username', 'невідомо'),
                               f"ВИДАЛИВ предмет: {subj_code} — {subj_name} (група ID {group_id})",
                               details=f"разом з {grade_count} оцінками" if grade_count > 0 else "оцінок не було")
                    flash(f'Предмет успешно удалён{"" if grade_count == 0 else f" вместе с {grade_count} оценками!"}', 'success')
            except Exception as e:
                conn.rollback()
                flash(f'Ошибка при удалении предмета: {str(e)}', 'error')

        elif action == 'move_up':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT position FROM subjects WHERE id=? AND group_id=?', (subject_id, group_id))
                current_position = cursor.fetchone()[0]
                cursor.execute('SELECT id, position FROM subjects WHERE position<? AND group_id=? ORDER BY position DESC LIMIT 1', (current_position, group_id))
                prev_subject = cursor.fetchone()
                if prev_subject:
                    cursor.execute('UPDATE subjects SET position=? WHERE id=? AND group_id=?', (prev_subject['position'], subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position=? WHERE id=? AND group_id=?', (current_position, prev_subject['id'], group_id))
                    conn.commit()
                    flash('Предмет перемещен вверх', 'success')
            except (KeyError, ValueError):
                flash('Ошибка при перемещении предмета', 'error')

        elif action == 'move_down':
            try:
                subject_id = request.form['subject_id']
                cursor.execute('SELECT position FROM subjects WHERE id=? AND group_id=?', (subject_id, group_id))
                current_position = cursor.fetchone()[0]
                cursor.execute('SELECT id, position FROM subjects WHERE position>? AND group_id=? ORDER BY position ASC LIMIT 1', (current_position, group_id))
                next_subject = cursor.fetchone()
                if next_subject:
                    cursor.execute('UPDATE subjects SET position=? WHERE id=? AND group_id=?', (next_subject['position'], subject_id, group_id))
                    cursor.execute('UPDATE subjects SET position=? WHERE id=? AND group_id=?', (current_position, next_subject['id'], group_id))
                    conn.commit()
                    flash('Предмет перемещен вниз', 'success')
            except (KeyError, ValueError):
                flash('Ошибка при перемещении предмета', 'error')

        elif action == 'edit_grades':
            try:
                subject_id = request.form['subject_id']
                subj_row = cursor.execute('SELECT name FROM subjects WHERE id=?', (subject_id,)).fetchone()
                subj_name = subj_row['name'] if subj_row else subject_id
                cursor.execute('SELECT id FROM students WHERE group_id=?', (group_id,))
                student_ids = [row['id'] for row in cursor.fetchall()]
                filled = 0
                for sid in student_ids:
                    grade = request.form.get(f'grade_{sid}')
                    grade_id = request.form.get(f'grade_id_{sid}')
                    if grade:
                        try:
                            grade = int(grade)
                            if not (0 <= grade <= 100):
                                flash(f'Оценка для студента {sid} должна быть от 0 до 100', 'error')
                                continue
                            if grade_id:
                                cursor.execute('UPDATE grades SET grade=? WHERE id=? AND student_id=? AND subject_id=?',
                                               (grade, grade_id, sid, subject_id))
                            else:
                                cursor.execute('INSERT INTO grades (student_id, subject_id, grade) VALUES (?, ?, ?)',
                                               (sid, subject_id, grade))
                            filled += 1
                        except ValueError:
                            flash(f'Некорректная оценка для студента {sid}', 'error')
                    else:
                        if grade_id:
                            cursor.execute('DELETE FROM grades WHERE id=? AND student_id=? AND subject_id=?',
                                           (grade_id, sid, subject_id))
                conn.commit()
                log_action(session.get('username', 'невідомо'),
                           f"змінив оцінки з предмету: {subj_name} (група ID {group_id})",
                           details=f"заповнено {filled} з {len(student_ids)} студентів")
                flash('Оценки обновлены', 'success')
            except (KeyError, ValueError) as e:
                flash(f'Ошибка при обновлении оценок: {str(e)}', 'error')

        conn.close()
        return redirect(url_for('admin.manage_subjects', group_id=group_id))

    conn.close()
    return render_template('admin_subjects.html', groups=groups, selected_group_id=selected_group_id,
                           subjects=subjects, students=students, grades=grades, selected_subject_id=selected_subject_id)


@admin_bp.route('/admin/manage_activities', methods=['GET', 'POST'])
@permission_required('manage_activities')
def manage_activities():
    conn = get_db()
    cursor = conn.cursor()

    cursor.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups WHERE archived = FALSE ORDER BY name, start_year
    """)
    groups = cursor.fetchall()

    selected_group_id = request.args.get('group_id')
    selected_entity_type = request.args.get('entity_type', 'practice')
    selected_entity_id = request.args.get('entity_id')
    entities = []
    students = []
    grades = []

    if selected_group_id:
        try:
            selected_group_id = int(selected_group_id)
            cursor.execute('SELECT id FROM groups WHERE id=?', (selected_group_id,))
            if not cursor.fetchone():
                flash('Обрана група не існує', 'error')
                selected_group_id = ''
            else:
                ALLOWED_TABLES = {'practice': 'practices', 'coursework': 'courseworks', 'attestation': 'attestations'}
                entity_table = ALLOWED_TABLES.get(selected_entity_type, 'practices')
                cursor.execute(f'SELECT * FROM {entity_table} WHERE group_id=? ORDER BY position', (selected_group_id,))
                entities = cursor.fetchall()

                if selected_entity_id:
                    try:
                        selected_entity_id = int(selected_entity_id)
                        cursor.execute('SELECT * FROM students WHERE group_id=?', (selected_group_id,))
                        students = cursor.fetchall()
                        import locale
                        locale.setlocale(locale.LC_COLLATE, 'uk_UA.UTF-8')
                        students = sorted(students, key=lambda s: locale.strxfrm(
                            f"{s['last_name_UA']} {s['first_name_UA']} {s['middle_name_UA']}"))
                        cursor.execute('SELECT id, student_id, entity_id, entity_type, grade, name FROM activity_grades WHERE entity_id=? AND entity_type=?',
                                       (selected_entity_id, selected_entity_type))
                        grades = cursor.fetchall()
                    except ValueError:
                        flash('Некоректний ID діяльності', 'error')
                        selected_entity_id = ''
        except ValueError:
            flash('Некоректний ID групи', 'error')
            selected_group_id = ''

    if request.method == 'POST':
        action = request.form.get('action')
        group_id = request.form.get('group_id')
        entity_type = request.form.get('entity_type', 'practice')
        ALLOWED_TABLES = {'practice': 'practices', 'coursework': 'courseworks', 'attestation': 'attestations'}
        if entity_type not in ALLOWED_TABLES:
            flash('Невірний тип діяльності', 'error')
            return redirect(url_for('admin.manage_activities'))
        entity_table = ALLOWED_TABLES[entity_type]

        try:
            if not group_id:
                flash('ID групи не вказано', 'error')
                return redirect(url_for('admin.manage_activities', group_id=selected_group_id, entity_type=entity_type))

            group_id = int(group_id)
            cursor.execute('SELECT id FROM groups WHERE id=?', (group_id,))
            if not cursor.fetchone():
                flash('Обрана група не існує', 'error')
                return redirect(url_for('admin.manage_activities', entity_type=entity_type))

            if action == 'add':
                code = request.form.get('code')
                name = request.form.get('name')
                credits = request.form.get('credits')
                type_ = request.form.get('type')
                position = request.form.get('position')

                if not all([code, name, credits, type_, position]) and entity_type != 'attestation':
                    flash('Усі поля мають бути заповнені', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))

                credits = int(credits) if credits else 0
                position = int(position) if position else 1
                if type_ not in ['Залік', 'Екзамен']:
                    flash('Невірний тип оцінки', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))

                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id=?', (group_id,))
                max_position = cursor.fetchone()[0] or 0
                if position <= max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position=position+1 WHERE position>=? AND group_id=?', (position, group_id))

                cursor.execute(f'INSERT INTO {entity_table} (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)',
                               (code, name or '', credits, type_, position, group_id))
                conn.commit()
                log_action(session.get('username', 'невідомо'),
                           f"додав {entity_type}: {code} — {name} (група ID {group_id})",
                           details=f"кредити: {credits}, тип: {type_}")
                flash('Діяльність додано', 'success')

            elif action == 'edit':
                entity_id = request.form.get('entity_id')
                if not entity_id:
                    flash('ID діяльності не вказано', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))

                entity_id = int(entity_id)
                code = request.form.get('code')
                name = request.form.get('name')
                credits = int(request.form.get('credits') or 0)
                type_ = request.form.get('type')
                position = int(request.form.get('position'))

                cursor.execute(f'SELECT id FROM {entity_table} WHERE id=? AND group_id=?', (entity_id, group_id))
                if not cursor.fetchone():
                    flash('Діяльність з вказаним ID не знайдено', 'error')
                    return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))

                cursor.execute(f'SELECT position FROM {entity_table} WHERE id=?', (entity_id,))
                current_position = cursor.fetchone()[0]
                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id=?', (group_id,))
                max_position = cursor.fetchone()[0] or 0
                if position != current_position and position <= max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position=position+1 WHERE position>=? AND group_id=? AND id!=?',
                                   (position, group_id, entity_id))

                cursor.execute(f'UPDATE {entity_table} SET code=?, name=?, credits=?, type=?, position=? WHERE id=? AND group_id=?',
                               (code, name or '', credits, type_, position, entity_id, group_id))
                conn.commit()
                log_action(session.get('username', 'невідомо'),
                           f"редагував {entity_type}: {code} — {name} (група ID {group_id})")
                flash('Діяльність оновлено', 'success')

            elif action == 'delete':
                entity_id = int(request.form.get('entity_id'))
                cursor.execute(f'SELECT position, code, name FROM {entity_table} WHERE id=?', (entity_id,))
                row = cursor.fetchone()
                position = row[0]
                cursor.execute(f'DELETE FROM {entity_table} WHERE id=? AND group_id=?', (entity_id, group_id))
                cursor.execute(f'UPDATE {entity_table} SET position=position-1 WHERE position>? AND group_id=?', (position, group_id))
                cursor.execute('DELETE FROM activity_grades WHERE entity_id=? AND entity_type=?', (entity_id, entity_type))
                conn.commit()
                log_action(session.get('username', 'невідомо'),
                           f"ВИДАЛИВ {entity_type}: {row[1]} — {row[2]} (група ID {group_id})")
                flash('Діяльність видалено', 'success')

            elif action == 'move_up':
                entity_id = int(request.form.get('entity_id'))
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id=?', (entity_id,))
                current_position = cursor.fetchone()[0]
                if current_position > 1:
                    cursor.execute(f'UPDATE {entity_table} SET position=? WHERE position=? AND group_id=?',
                                   (current_position, current_position - 1, group_id))
                    cursor.execute(f'UPDATE {entity_table} SET position=? WHERE id=? AND group_id=?',
                                   (current_position - 1, entity_id, group_id))
                    conn.commit()
                    flash('Діяльність переміщено вгору', 'success')

            elif action == 'move_down':
                entity_id = int(request.form.get('entity_id'))
                cursor.execute(f'SELECT position FROM {entity_table} WHERE id=?', (entity_id,))
                current_position = cursor.fetchone()[0]
                cursor.execute(f'SELECT MAX(position) FROM {entity_table} WHERE group_id=?', (group_id,))
                max_position = cursor.fetchone()[0]
                if current_position < max_position:
                    cursor.execute(f'UPDATE {entity_table} SET position=? WHERE position=? AND group_id=?',
                                   (current_position, current_position + 1, group_id))
                    cursor.execute(f'UPDATE {entity_table} SET position=? WHERE id=? AND group_id=?',
                                   (current_position + 1, entity_id, group_id))
                    conn.commit()
                    flash('Діяльність переміщено вниз', 'success')

            elif action == 'edit_grades':
                entity_id = request.form['entity_id']
                entity_row = cursor.execute(f'SELECT name FROM {entity_table} WHERE id=?', (entity_id,)).fetchone()
                entity_name = entity_row['name'] if entity_row else entity_id
                cursor.execute('SELECT id FROM students WHERE group_id=?', (group_id,))
                student_ids = [row['id'] for row in cursor.fetchall()]
                filled = 0
                for sid in student_ids:
                    grade = request.form.get(f'grade_{sid}')
                    grade_id = request.form.get(f'grade_id_{sid}')
                    name_val = request.form.get(f'name_{sid}', '') if entity_type == 'attestation' else ''

                    if grade or name_val:
                        try:
                            grade_int = int(grade) if grade else None
                            if grade_int is not None and not (0 <= grade_int <= 100):
                                flash(f'Оценка для студента {sid} должна быть от 0 до 100', 'error')
                                continue
                            if grade_id:
                                cursor.execute('UPDATE activity_grades SET grade=?, name=? WHERE id=? AND student_id=? AND entity_id=? AND entity_type=?',
                                               (grade_int, name_val, grade_id, sid, entity_id, entity_type))
                            else:
                                cursor.execute('INSERT INTO activity_grades (student_id, entity_id, entity_type, grade, name) VALUES (?, ?, ?, ?, ?)',
                                               (sid, entity_id, entity_type, grade_int, name_val))
                            filled += 1
                        except ValueError:
                            flash(f'Некорректная оценка для студента {sid}', 'error')
                    else:
                        if grade_id:
                            cursor.execute('DELETE FROM activity_grades WHERE id=? AND student_id=? AND entity_id=? AND entity_type=?',
                                           (grade_id, sid, entity_id, entity_type))
                conn.commit()
                log_action(session.get('username', 'невідомо'),
                           f"змінив оцінки з {entity_type}: {entity_name} (група ID {group_id})",
                           details=f"заповнено {filled} з {len(student_ids)} студентів")
                flash('Оценки обновлены', 'success')

            return redirect(url_for('admin.manage_activities', group_id=group_id, entity_type=entity_type))

        except (ValueError, sqlite3.Error) as e:
            conn.rollback()
            flash(f'Помилка: {e}', 'error')
            return redirect(url_for('admin.manage_activities', group_id=selected_group_id, entity_type=entity_type))

    conn.close()
    return render_template('admin_activities.html', groups=groups, selected_group_id=selected_group_id,
                           entities=entities, students=students, grades=grades,
                           selected_entity_id=selected_entity_id, entity_type=selected_entity_type)


@admin_bp.route('/admin/view_logs')
@permission_required('view_logs')
def view_logs():
    current_dir = os.path.dirname(__file__)
    project_root = os.path.dirname(current_dir)
    log_file_path = os.path.join(project_root, 'app.log')

    parsed_logs = []

    if os.path.exists(log_file_path):
        try:
            with open(log_file_path, 'r', encoding='utf-8') as f:
                for line in f:
                    line = line.strip()
                    if not line:
                        continue
                    entry = {'raw': line, 'date': '', 'time': '', 'level': 'INFO', 'username': '', 'action': line}
                    m = re.match(r'(\d{4}-\d{2}-\d{2})\s*\|\s*(\d{2}:\d{2}:\d{2})\s*\|\s*(\w+)\s*\|\s*(.*)', line)
                    if m:
                        entry['date'] = m.group(1)
                        entry['time'] = m.group(2)
                        entry['level'] = m.group(3)
                        rest = m.group(4).strip()
                        entry['action'] = rest
                        u = re.match(r'👤\s*([^\s-][^-]*?)\s+-\s+(.*)', rest)
                        if u:
                            entry['username'] = u.group(1).strip()
                            entry['action'] = u.group(2).strip()
                    parsed_logs.append(entry)
        except Exception as e:
            logging.error(f"Помилка при читанні логів: {e}")

    parsed_logs.reverse()
    usernames = sorted({e['username'] for e in parsed_logs if e['username']})

    log_action(session.get('username', 'невідомо'), "переглянув журнал дій")

    from datetime import date
    return render_template('view_logs.html', logs=parsed_logs, usernames=usernames,
                           now=date.today().strftime('%Y-%m-%d'))


@admin_bp.route('/admin/users', methods=['GET', 'POST'])
@permission_required('manage_users')
def manage_users():
    conn = get_db()
    try:
        users = conn.execute("""
            SELECT u.id, u.username, u.role, u.is_admin, u.permissions,
                   GROUP_CONCAT(g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)', ', ') AS group_names
            FROM users u
            LEFT JOIN user_groups ug ON u.id = ug.user_id
            LEFT JOIN groups g ON ug.group_id = g.id
            GROUP BY u.id ORDER BY u.username
        """).fetchall()

        perm_names_ua = {
            'manage_users': 'Список користувачів та управління правами',
            'view_logs': 'Журнал дій',
            'group_export': 'Масова генерація документів',
            'import_from_excel': 'Інпорт студентів',
            'manage_education_documents': 'Управління документами про освіту',
            'study_periods': 'Періоди навчання',
            'manage_groups': 'Управління групами',
            'manage_subjects': 'Предмети',
            'manage_activities': 'Управління діяльностями',
            'import_subjects': 'Імпорт предметів з Excel',
            'archive': 'Управління архівом',
            'manage_students': 'Управління студентами (Видалення студента та його війс. док.)',
            'manage_accreditations': 'Управління акредетаціями',
            'manage_diplomas': 'Управління номерами диплому і додатку',
            'import_education_docs': 'Управління імпортом документів'
        }

        if request.method == 'POST':
            user_id = request.form.get('user_id')
            if not user_id:
                flash('Не вказано користувача', 'danger')
                return redirect(url_for('admin.manage_users'))

            is_admin = 1 if 'is_admin' in request.form else 0
            selected_perms = [p for p in PERMISSIONS if p in request.form]

            target_user = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()

            conn.execute("UPDATE users SET is_admin=?, permissions=? WHERE id=?",
                         (is_admin, json.dumps(selected_perms), user_id))
            conn.commit()

            log_action(
                session.get('username', 'невідомо'),
                f"змінив права: {target_user['username']} (ID {user_id})",
                details=f"is_admin: {bool(is_admin)}, дозволи: {', '.join(selected_perms) or 'жодного'}"
            )
            flash('Права успішно оновлено', 'success')
            return redirect(url_for('admin.manage_users'))

        return render_template('manage_users.html', users=users, permissions=PERMISSIONS, perm_names_ua=perm_names_ua)

    except sqlite3.Error as e:
        flash(f'Помилка бази даних: {e}', 'danger')
        return redirect(url_for('admin.manage_users'))
    finally:
        conn.close()


@admin_bp.route('/admin/users/add', methods=['GET', 'POST'])
@permission_required('manage_users')
def add_user():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        role = request.form.get('role')
        group_ids = request.form.getlist('group_id')

        if not all([username, password, role]):
            flash('Заповніть усі обовязкові поля', 'danger')
            return redirect(url_for('admin.add_user'))

        is_admin = 1 if role == 'admin' else 0
        permissions = json.dumps([])

        conn = get_db()
        try:
            exists = conn.execute("SELECT 1 FROM users WHERE username=?", (username,)).fetchone()
            if exists:
                flash(f'Користувач з іменем "{username}" вже існує', 'danger')
                return redirect(url_for('admin.add_user'))

            cursor = conn.execute(
                "INSERT INTO users (username, password_hash, role, is_admin, permissions) VALUES (?, ?, ?, ?, ?)",
                (username, generate_password_hash(password), role, is_admin, permissions)
            )

            user_id = cursor.lastrowid

            for gid in group_ids:
                if gid:
                    conn.execute("INSERT INTO user_groups (user_id, group_id) VALUES (?, ?)", (user_id, gid))

            conn.commit()
            log_action(
                session.get('username', 'невідомо'),
                f"додав користувача: {username} (ID {user_id})",
                details=f"роль: {role}, груп: {len(group_ids)}"
            )
            flash('Користувача успішно додано', 'success')
            return redirect(url_for('admin.manage_users'))

        except sqlite3.IntegrityError as e:
            flash(f'Помилка бази даних: {e}', 'danger')
        finally:
            conn.close()

    conn = get_db()
    try:
        groups = conn.execute("""
            SELECT id, name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
            FROM groups ORDER BY name, start_year
        """).fetchall()
    finally:
        conn.close()

    return render_template('add_user.html', groups=groups)


@admin_bp.route('/admin/users/<int:user_id>/edit', methods=['GET', 'POST'])
@permission_required('manage_users')
def edit_user(user_id):
    conn = get_db()
    try:
        if request.method == 'POST':
            role = request.form.get('role')
            group_ids = request.form.getlist('group_id')

            if not role:
                flash('Роль обовязкова', 'danger')
                return redirect(url_for('admin.edit_user', user_id=user_id))

            is_admin = 1 if role == 'admin' else 0
            user_row = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()

            conn.execute("UPDATE users SET role=?, is_admin=? WHERE id=?", (role, is_admin, user_id))
            conn.execute("DELETE FROM user_groups WHERE user_id=?", (user_id,))
            for gid in group_ids:
                if gid:
                    conn.execute("INSERT INTO user_groups (user_id, group_id) VALUES (?, ?)", (user_id, gid))

            conn.commit()
            log_action(
                session.get('username', 'невідомо'),
                f"змінив роль/групи: {user_row['username']} (ID {user_id})",
                details=f"роль: {role}, груп: {len(group_ids)}"
            )
            flash('Дані користувача оновлено', 'success')
            return redirect(url_for('admin.manage_users'))

        user = conn.execute("SELECT id, username, role FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            flash('Користувача не знайдено', 'danger')
            return redirect(url_for('admin.manage_users'))

        current_groups = conn.execute("SELECT group_id FROM user_groups WHERE user_id=?", (user_id,)).fetchall()
        current_group_ids = [row['group_id'] for row in current_groups]

        groups = conn.execute("""
            SELECT id, name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
            FROM groups ORDER BY name, start_year
        """).fetchall()

        return render_template('edit_user.html', user=user, groups=groups, current_group_ids=current_group_ids)
    finally:
        conn.close()


@admin_bp.route('/admin/users/<int:user_id>/change-password', methods=['GET', 'POST'])
@permission_required('manage_users')
def change_password(user_id):
    if request.method == 'POST':
        password = request.form.get('password')
        if not password or len(password) < 6:
            flash('Пароль повинен бути не коротшим 6 символів', 'danger')
            return redirect(url_for('admin.change_password', user_id=user_id))

        conn = get_db()
        try:
            target = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
            conn.execute("UPDATE users SET password_hash=? WHERE id=?", (generate_password_hash(password), user_id))
            conn.commit()
            log_action(
                session.get('username', 'невідомо'),
                f"змінив пароль: {target['username']} (ID {user_id})"
            )
            flash('Пароль успішно змінено', 'success')
            return redirect(url_for('admin.manage_users'))
        finally:
            conn.close()

    conn = get_db()
    try:
        user = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
        if not user:
            flash('Користувача не знайдено', 'danger')
            return redirect(url_for('admin.manage_users'))
    finally:
        conn.close()

    return render_template('change_password.html', user_id=user_id, username=user['username'])


@admin_bp.route('/admin/users/<int:user_id>/delete', methods=['POST'])
@permission_required('manage_users')
def delete_user(user_id):
    conn = get_db()
    try:
        username_row = conn.execute("SELECT username FROM users WHERE id=?", (user_id,)).fetchone()
        if not username_row:
            flash('Користувача не знайдено', 'danger')
            return redirect(url_for('admin.manage_users'))

        conn.execute("DELETE FROM users WHERE id=?", (user_id,))
        conn.execute("DELETE FROM user_groups WHERE user_id=?", (user_id,))
        conn.commit()

        log_action(
            session.get('username', 'невідомо'),
            f"ВИДАЛИВ користувача: {username_row['username']} (ID {user_id})"
        )
        flash('Користувача успішно видалено', 'success')
    except sqlite3.Error as e:
        flash(f'Помилка при видаленні: {e}', 'danger')
    finally:
        conn.close()

    return redirect(url_for('admin.manage_users'))


@admin_bp.route('/admin/group_export', methods=['GET', 'POST'])
@permission_required('group_export')
def group_export():
    conn = get_db()
    conn.row_factory = sqlite3.Row

    groups = conn.execute("""
        SELECT id, name, start_year, study_form, program_credits,
               name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
        FROM groups WHERE archived = FALSE ORDER BY name, start_year
    """).fetchall()

    available_templates = get_available_templates()
    default_template = available_templates[0] if available_templates else ''

    current_year = datetime.now().year
    years = list(range(1980, current_year + 1))
    students = []
    selected_group_id = request.args.get('group_id', type=int) if request.method == 'GET' else request.form.get('group_id', type=int)
    selected_year = request.args.get('birth_year', type=int) if request.method == 'GET' else request.form.get('birth_year', type=int)
    selected_template = request.args.get('template', default_template) if request.method == 'GET' else request.form.get('template', default_template)

    if selected_group_id:
        group_check = conn.execute("SELECT id FROM groups WHERE id=? AND archived=FALSE", (selected_group_id,)).fetchone()
        if not group_check:
            flash('Обрана група не існує або є архівною.', 'error')
            selected_group_id = None

    if request.method == 'POST':
        if not selected_group_id and not selected_year:
            flash('Будь ласка, оберіть групу або рік народження.', 'error')
        else:
            active_students = request.form.getlist('active_students')
            return redirect(url_for('admin.generate_group_docs', group_id=selected_group_id,
                                    birth_year=selected_year, template=selected_template,
                                    active_students=','.join(active_students)))

    if selected_group_id or selected_year:
        base_query = "SELECT * FROM students WHERE archived = FALSE"
        params = []
        if selected_group_id:
            base_query += " AND group_id=?"
            params.append(selected_group_id)
        if selected_year:
            base_query += " AND SUBSTR(birth_date, 7, 4) >= ?"
            params.append(str(selected_year))
        try:
            students = conn.execute(base_query, params).fetchall()
        except Exception as e:
            logging.error(f"Помилка при отриманні студентів: {e}")
            conn.close()
            return "Помилка бази даних", 500

    conn.close()
    return render_template('group_export.html', students=students, groups=groups, years=years,
                           selected_group_id=selected_group_id, selected_year=selected_year,
                           selected_template=selected_template,
                           available_templates=available_templates)


UPLOAD_FOLDER = 'Uploads'
ALLOWED_EXTENSIONS = {'xlsx'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@admin_bp.route('/admin/import_subjects', methods=['GET', 'POST'])
@permission_required('import_subjects')
def import_subjects():
    conn = get_db()
    cursor = conn.cursor()

    try:
        cursor.execute("""
            SELECT id, name, start_year, study_form, program_credits,
                   name || ' (' || start_year || ', ' || study_form || ', ' || program_credits || ' кредитів)' AS display_name
            FROM groups WHERE archived = FALSE ORDER BY name, start_year
        """)
        groups = cursor.fetchall()
        if not groups:
            flash("Немає доступних груп", "warning")
    except sqlite3.Error as e:
        logging.error(f"Database error while fetching groups: {e}")
        flash("Помилка бази даних при отриманні груп", "error")
        groups = []

    selected_group_id = request.args.get('group_id', '')

    if request.method == 'POST':
        file = request.files.get('excel_file')
        group_id = request.form.get('group_id')

        if not file:
            flash("Будь ласка, виберіть файл", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

        ext = file.filename.lower().split('.')[-1]
        if ext not in ['xlsx', 'xlsm', 'xltx', 'xltm']:
            flash("❗ Підтримуються тільки Excel файли формату .xlsx", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

        if not group_id:
            flash("ID групи не вказано", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

        try:
            group_id = int(group_id)
            cursor.execute("SELECT id FROM groups WHERE id=?", (group_id,))
            if not cursor.fetchone():
                flash("Обрана група не існує", "error")
                return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)
        except ValueError:
            flash("Некоректний ID групи", "error")
            return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

        filename = f"subjects_{int(time.time())}.xlsx"
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        os.makedirs(UPLOAD_FOLDER, exist_ok=True)
        file.save(filepath)

        try:
            try:
                wb = openpyxl.load_workbook(filepath, data_only=True)
                sheet = wb.active
            except InvalidFileException:
                flash("❗ Файл має неправильний формат. Збережіть його як Excel (*.xlsx)", "error")
                os.remove(filepath)
                return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)

            inserted = 0
            skipped = 0
            cursor.execute("SELECT MAX(position) FROM subjects WHERE group_id=?", (group_id,))
            max_position = cursor.fetchone()[0] or 0
            current_position = max_position + 1

            for i, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                if not row or all(cell is None for cell in row):
                    continue
                try:
                    code, name, credits, type_ = row
                    if not all([code, name, credits, type_]):
                        skipped += 1
                        continue
                    code = str(code).strip()
                    name = str(name).strip()
                    type_ = str(type_).strip()
                    if type_ not in ['Залік', 'Екзамен']:
                        flash(f"❗ Невірний тип у рядку {i}: {type_}", "error")
                        skipped += 1
                        continue
                    credits = int(credits)
                    if credits < 1:
                        flash(f"❗ Некоректні кредити у рядку {i}", "error")
                        skipped += 1
                        continue
                    cursor.execute("SELECT id FROM subjects WHERE group_id=? AND code=?", (group_id, code))
                    if cursor.fetchone():
                        skipped += 1
                        continue
                    cursor.execute("INSERT INTO subjects (code, name, credits, type, position, group_id) VALUES (?, ?, ?, ?, ?, ?)",
                                   (code, name, credits, type_, current_position, group_id))
                    inserted += 1
                    current_position += 1
                except Exception as e:
                    logging.error(f"Row {i} error: {e}")
                    skipped += 1
                    continue

            conn.commit()
            log_action(
                session.get('username', 'невідомо'),
                f"імпорт предметів з Excel: додано {inserted}, пропущено {skipped}",
                details=f"група ID: {group_id}, файл: {filename}"
            )
            flash(f"✅ Імпорт завершено. Додано: {inserted}, пропущено: {skipped}", "success")

        except Exception as e:
            conn.rollback()
            flash(f"⚠️ Помилка при імпорті Excel: {e}", "error")
            logging.error(f"Error importing Excel for group_id={group_id}: {e}")
        finally:
            if os.path.exists(filepath):
                os.remove(filepath)
            conn.close()

        return redirect(url_for('admin.manage_subjects', group_id=group_id))

    conn.close()
    return render_template('import_subjects.html', groups=groups, selected_group_id=selected_group_id)


@admin_bp.route('/admin/generate_group_docs', methods=['GET', 'POST'])
@permission_required('group_export')
def generate_group_docs():
    group_id = request.args.get('group_id', type=int) if request.method == 'GET' else request.form.get('group_id', type=int)
    birth_year = request.args.get('birth_year', type=int) if request.method == 'GET' else request.form.get('birth_year', type=int)
    selected_template = request.args.get('template', '') if request.method == 'GET' else request.form.get('template', '')
    active_students = request.args.get('active_students', '').split(',') if request.args.get('active_students') else []

    if not group_id and not birth_year:
        flash('Оберіть групу або рік народження для генерації документів.', 'error')
        return redirect(url_for('admin.group_export'))

    conn = get_db()
    conn.row_factory = sqlite3.Row
    base_query = """
        SELECT s.*,
               g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS group_name,
               g.study_form, g.start_year, g.program_credits,
               g.qualification_name, g.degree_level, g.specialty, g.educational_program, g.knowledge_area,
               g.qualification_name_en, g.degree_level_en, g.specialty_en, g.educational_program_en, g.knowledge_area_en,
               g.institution_name_and_status, g.institution_name_and_status_en,
               g.entry_requirements, g.entry_requirements_en,
               g.learning_outcomes, g.learning_outcomes_en, g.program_includes, g.program_includes_en
        FROM students s LEFT JOIN groups g ON s.group_id = g.id WHERE s.archived = FALSE
    """
    params = []
    if group_id:
        base_query += " AND s.group_id=?"
        params.append(group_id)
    if birth_year:
        base_query += " AND SUBSTR(s.birth_date, 7, 4) >= ?"
        params.append(str(birth_year))

    try:
        students = conn.execute(base_query, params).fetchall()
        if not students:
            conn.close()
            return "Студенты не найдены по заданным фильтрам", 404
    except Exception as e:
        logging.error(f"Ошибка при выполнении SQL-запроса: {e}")
        conn.close()
        return "Ошибка базы данных", 500

    if active_students and active_students[0]:
        students = [s for s in students if str(s['id']) in active_students]

    group_name = "Зі всіх груп"
    if group_id and students:
        group_name = students[0]['group_name'] if students[0]['group_name'] else f"Група_{group_id}"

    output_dir = os.path.join(os.getcwd(), 'generated_docs')
    os.makedirs(output_dir, exist_ok=True)
    zip_filename = f"{group_name}_{str(birth_year) if birth_year else 'Всі роки народження'}.zip"
    zip_path = os.path.join(output_dir, zip_filename)

    try:
        with ZipFile(zip_path, 'w') as zipf:
            for student in students:
                student_dict = dict(student)
                military = conn.execute("SELECT * FROM military WHERE student_id=?", (student['id'],)).fetchone()
                military_dict = dict(military) if military else {}
                filename = f"{student_dict['last_name_UA']}_{student_dict['first_name_UA']}.docx".replace(" ", "_")
                full_path = os.path.join(output_dir, filename)
                try:
                    gen_doc(student_dict, military_dict, template=selected_template, out=full_path,
                            user_name=session.get('username', 'невідомо'))
                    zipf.write(full_path, arcname=filename)
                except Exception as e:
                    logging.error(f"Ошибка при генерации документа для {student_dict.get('last_name_UA', '')}: {e}")
                    continue

        log_action(
            session.get('username', 'невідомо'),
            f"масова генерація документів: {group_name}",
            details=f"шаблон: {selected_template}, рік нар.: {birth_year or 'всі'}, студентів: {len(students)}"
        )
    finally:
        conn.close()

    return send_file(zip_path, as_attachment=True)


@admin_bp.route('/admin/archive/<int:group_id>', methods=['POST'])
@permission_required('archive')
def archive_group(group_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, start_year FROM groups WHERE id=?", (group_id,))
    group_row = cursor.fetchone()
    if not group_row:
        flash('Група не знайдена', 'error')
        conn.close()
        return redirect(url_for('admin.manage_groups'))

    cursor.execute("UPDATE groups SET archived = TRUE WHERE id=?", (group_id,))
    cursor.execute("UPDATE students SET archived = TRUE WHERE group_id=?", (group_id,))
    conn.commit()
    conn.close()

    log_action(
        session.get('username', 'невідомо'),
        f"заархівував групу: {group_row['name']} ({group_row['start_year']}) (ID {group_id})"
    )
    flash('Групу успішно заархівовано', 'success')
    return redirect(url_for('admin.manage_groups'))


@admin_bp.route('/admin/unarchive_group/<int:group_id>', methods=['POST'])
@permission_required('archive')
def unarchive_group(group_id):
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, start_year FROM groups WHERE id=? AND archived=TRUE", (group_id,))
    group_row = cursor.fetchone()
    if not group_row:
        flash('Архівна група не знайдена', 'error')
        conn.close()
        return redirect(url_for('admin.archive'))

    cursor.execute("UPDATE groups SET archived = FALSE WHERE id=?", (group_id,))
    cursor.execute("UPDATE students SET archived = FALSE WHERE group_id=?", (group_id,))
    conn.commit()
    conn.close()

    log_action(
        session.get('username', 'невідомо'),
        f"розархівував групу: {group_row['name']} ({group_row['start_year']}) (ID {group_id})"
    )
    flash('Групу успішно розархівовано', 'success')
    return redirect(url_for('admin.archive'))


@admin_bp.route('/admin/archive')
@permission_required('archive')
def archive():
    conn = get_db()
    conn.row_factory = sqlite3.Row

    groups = conn.execute("""
        SELECT g.id, g.name, g.start_year, g.study_form, g.program_credits,
               g.name || ' (' || g.start_year || ', ' || g.study_form || ', ' || g.program_credits || ' кредитів)' AS display_name,
               (SELECT COUNT(*) FROM students s WHERE s.group_id = g.id AND s.archived = TRUE) AS student_count
        FROM groups g WHERE g.archived = TRUE ORDER BY g.start_year DESC, g.name
    """).fetchall()

    students_by_group = {}
    for group in groups:
        students = conn.execute("""
            SELECT id, last_name_UA, first_name_UA, birth_date FROM students
            WHERE group_id=? AND archived=TRUE ORDER BY last_name_UA
        """, (group['id'],)).fetchall()
        students_by_group[group['id']] = students

    conn.close()
    log_action(session.get('username', 'невідомо'), "переглянув список архівних груп")
    return render_template('archive.html', groups=groups, students_by_group=students_by_group)


TEMP_PREVIEW_FOLDER = "temp_preview"

def save_preview_to_file(preview):
    os.makedirs(TEMP_PREVIEW_FOLDER, exist_ok=True)
    preview_id = str(uuid.uuid4())
    path = os.path.join(TEMP_PREVIEW_FOLDER, f"{preview_id}.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(preview, f, ensure_ascii=False)
    return preview_id

def load_preview_from_file(preview_id):
    path = os.path.join(TEMP_PREVIEW_FOLDER, f"{preview_id}.json")
    if not os.path.exists(path):
        return []
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)

def fuzzy_find_student(cursor, full_name, threshold=80):
    cursor.execute("SELECT id, last_name_UA, first_name_UA, middle_name_UA FROM students WHERE archived=FALSE")
    students = cursor.fetchall()
    names = ["{} {} {}".format(s["last_name_UA"], s["first_name_UA"], s["middle_name_UA"]) for s in students]
    matches = process.extract(full_name, names, scorer=fuzz.ratio, limit=3)
    for match_name, score, idx in matches:
        if score >= threshold:
            return students[idx]["id"], match_name, score
    return None, None, None

def translate_to_en(text):
    if not text:
        return ""
    if text in translation_cache:
        return translation_cache[text]
    try:
        result = translator.translate(text)
        translation_cache[text] = result
        return result
    except Exception as e:
        logging.warning(f"Помилка перекладу '{text}': {e}")
        return text
        
def find_country(text):
    countries = {
        "польща": "Poland",
        "poland": "Poland",
        "німеччина": "Germany",
        "germany": "Germany",
        "чех": "Czech Republic",
        "czech": "Czech Republic",
        "словач": "Slovakia",
        "slovakia": "Slovakia"
    }

    lower = text.lower()

    for key, en in countries.items():
        if key in lower:
            return key.capitalize(), en

    if "україна " in lower:
        return "Україна ", "Ukraine"

    return "Україна", "Ukraine"

def parse_document(text: str):
    text = text.strip()
    pattern = re.compile(
        r'^(?P<type>[^;]+?)\s*;\s*(?P<date>\d{2}\.\d{2}\.\d{4})\s*;\s*Ким видано:\s*(?P<institution>.+?)$',
        re.IGNORECASE | re.UNICODE
    )
    match = pattern.search(text)
    if not match:
        return None

    full_prefix = match.group("type").strip()
    parts = re.split(r'\s{2,}', full_prefix.strip())
    if len(parts) < 2:
        doc_type = full_prefix
        doc_number = ""
    else:
        doc_type = " ".join(parts[:-1]).strip()
        doc_number = parts[-1].strip()

    completion_date = match.group("date").strip()
    completion_date = completion_date.replace(".", "/")
    institution = match.group("institution")
    country, country_en = find_country(institution)

    return {
        "document_type": doc_type,
        "document_type_en": translate_to_en(doc_type) or "",
        "document_number": doc_number,
        "completion_date": completion_date,
        "institution_name": institution,
        "institution_name_en": translate_to_en(institution) or "",
        "country": country,
        "country_en": country_en
    }

def parse_reference_cell_ua(text):
    if not text or not str(text).strip():
        return {
            "reference_number": "",
            "reference_institution": "",
            "reference_institution_en": "",
            "reference_country": "",
            "reference_country_en": "",
            "reference_issue_date": "",
        }
    parts = [p.strip() for p in str(text).split(";")]
    parts += [""] * (4 - len(parts))
    reference_number, reference_institution, reference_country, reference_issue_date = parts[:4]
    reference_issue_date = reference_issue_date.replace(".", "/")
    return {
        "reference_number": reference_number,
        "reference_institution": reference_institution,
        "reference_institution_en": translate_to_en(reference_institution) or "",
        "reference_country": reference_country,
        "reference_country_en": translate_to_en(reference_country) or "",
        "reference_issue_date": reference_issue_date,
    }

def parse_recognition_cell_ua(text):
    if not text or not str(text).strip():
        return {
            "recognition_certificate_number": "",
            "recognition_issuer": "",
            "recognition_issuer_en": "",
            "recognition_date": "",
        }
    parts = [p.strip() for p in str(text).split(";")]
    parts += [""] * (3 - len(parts))
    recognition_certificate_number, recognition_issuer, recognition_date = parts[:3]
    recognition_date = recognition_date.replace(".", "/")
    return {
        "recognition_certificate_number": recognition_certificate_number,
        "recognition_issuer": recognition_issuer,
        "recognition_issuer_en": translate_to_en(recognition_issuer) or "",
        "recognition_date": recognition_date,
    }

def import_documents_preview(file_path, db):
    wb = load_workbook(file_path)
    sheet = wb.active
    cursor = db.cursor()
    preview_rows = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        fio = row[0]
        document_text = row[1]
        reference_ua_text = row[2] if len(row) > 2 else None
        recognition_ua_text = row[3] if len(row) > 3 else None

        if not fio or not document_text:
            preview_rows.append({"row_index": row_index, "error": "Пусті дані"})
            continue
        student_id, matched_name, score = fuzzy_find_student(cursor, fio)
        if not student_id:
            preview_rows.append({"row_index": row_index, "error": f"Студент не знайдений: {fio}"})
            continue
        data = parse_document(document_text)
        if not data:
            preview_rows.append({"row_index": row_index, "error": f"Не вдалося розпізнати документ: {document_text}"})
            continue

        data.update(parse_reference_cell_ua(reference_ua_text))
        data.update(parse_recognition_cell_ua(recognition_ua_text))


        cursor.execute("""
            SELECT id, document_type, document_type_en, document_number, completion_date,
                   institution_name, institution_name_en, country, country_en
            FROM education_documents WHERE student_id=?
        """, (student_id,))
        existing_doc = cursor.fetchone()

        row_data = {"row_index": row_index, "student_id": student_id, "matched_name": matched_name, "score": score, **data}

        if existing_doc:
            row_data.update({
                "status": "Оновлення", "status_class": "text-warning",
                "existing_info": f"(існує: {existing_doc['document_number'] or 'без номера'})",
                "has_document": True,
                "existing_doc_id": existing_doc["id"],
                "old_document_type": existing_doc["document_type"] or "",
                "old_document_type_en": existing_doc["document_type_en"] or "",
                "old_document_number": existing_doc["document_number"] or "",
                "old_completion_date": existing_doc["completion_date"] or "",
                "old_institution_name": existing_doc["institution_name"] or "",
                "old_institution_name_en": existing_doc["institution_name_en"] or "",
                "old_country": existing_doc["country"] or "",
                "old_country_en": existing_doc["country_en"] or "",
            })

            cursor.execute("""
                SELECT reference_number, reference_institution, reference_institution_en,
                       reference_country, reference_country_en, reference_issue_date,
                       recognition_certificate_number, recognition_issuer, recognition_issuer_en, recognition_date
                FROM foreign_education_docs WHERE education_doc_id=?
            """, (existing_doc["id"],))
            existing_foreign = cursor.fetchone()
            if existing_foreign:
                row_data.update({
                    "old_reference_number": existing_foreign["reference_number"] or "",
                    "old_reference_institution": existing_foreign["reference_institution"] or "",
                    "old_reference_institution_en": existing_foreign["reference_institution_en"] or "",
                    "old_reference_country": existing_foreign["reference_country"] or "",
                    "old_reference_country_en": existing_foreign["reference_country_en"] or "",
                    "old_reference_issue_date": existing_foreign["reference_issue_date"] or "",
                    "old_recognition_certificate_number": existing_foreign["recognition_certificate_number"] or "",
                    "old_recognition_issuer": existing_foreign["recognition_issuer"] or "",
                    "old_recognition_issuer_en": existing_foreign["recognition_issuer_en"] or "",
                    "old_recognition_date": existing_foreign["recognition_date"] or "",
                })
        else:
            row_data.update({"status": "Новий", "status_class": "text-success", "existing_info": "", "has_document": False})

        preview_rows.append(row_data)
    return preview_rows
    
@admin_bp.route('/admin/import_education_docs_preview', methods=['GET', 'POST'])
@permission_required('import_education_docs')
def import_docs_preview():
    if request.method == "POST":
        file = request.files.get("file")
        if not file:
            flash("Файл не обрано", "danger")
            return redirect(url_for('admin.manage_education_documents'))

        filename = secure_filename(file.filename)
        path = os.path.join("uploads", filename)
        os.makedirs("uploads", exist_ok=True)
        file.save(path)

        db = get_db()
        preview = import_documents_preview(path, db)
        preview_id = save_preview_to_file(preview)

        return render_template("import_education_preview.html", preview=preview, preview_id=preview_id)

    return render_template("import_education_upload.html")

@admin_bp.route('/admin/import_docs_commit', methods=['POST'])
@permission_required('import_education_docs')
def import_docs_commit():
    preview_id = request.form.get("preview_id")
    preview = load_preview_from_file(preview_id)
    db = get_db()
    added = 0
    updated = 0

    for row in preview:
        row_index = row["row_index"]
        if f"add_{row_index}" not in request.form:
            continue

        if not row.get("error"):
            row["document_type"]       = request.form.get(f"document_type_{row_index}",       row["document_type"])
            row["document_type_en"]    = request.form.get(f"document_type_en_{row_index}",    row.get("document_type_en", "")) or ""
            row["document_number"]     = request.form.get(f"document_number_{row_index}",     row["document_number"])
            row["completion_date"]     = request.form.get(f"completion_date_{row_index}",     row["completion_date"])
            row["institution_name"]    = request.form.get(f"institution_name_{row_index}",    row["institution_name"])
            row["institution_name_en"] = request.form.get(f"institution_name_en_{row_index}", row.get("institution_name_en", "")) or ""
            row["country"]             = request.form.get(f"country_{row_index}",             row["country"])
            row["country_en"]          = request.form.get(f"country_en_{row_index}",          row["country_en"])

            reference_number = request.form.get(f"reference_number_{row_index}", row.get("reference_number", "")) or None
            reference_institution = request.form.get(f"reference_institution_{row_index}", row.get("reference_institution", "")) or None
            reference_institution_en = request.form.get(f"reference_institution_en_{row_index}", row.get("reference_institution_en", "")) or None
            reference_country = request.form.get(f"reference_country_{row_index}", row.get("reference_country", "")) or None
            reference_country_en = request.form.get(f"reference_country_en_{row_index}", row.get("reference_country_en", "")) or None
            reference_issue_date = request.form.get(f"reference_issue_date_{row_index}", row.get("reference_issue_date", "")) or None
            recognition_certificate_number = request.form.get(f"recognition_certificate_number_{row_index}", row.get("recognition_certificate_number", "")) or None
            recognition_issuer = request.form.get(f"recognition_issuer_{row_index}", row.get("recognition_issuer", "")) or None
            recognition_issuer_en = request.form.get(f"recognition_issuer_en_{row_index}", row.get("recognition_issuer_en", "")) or None
            recognition_date = request.form.get(f"recognition_date_{row_index}", row.get("recognition_date", "")) or None

            country_lower = (row["country"] or "").lower()

            is_foreign = not (country_lower in ('україна', 'ukraine'))

            cursor = db.cursor()
            cursor.execute("SELECT id FROM education_documents WHERE student_id=?", (row["student_id"],))
            existing = cursor.fetchone()

            doc_type_en  = row.get("document_type_en", "") or ""
            inst_name_en = row.get("institution_name_en", "") or ""
            country_en   = row.get("country_en", "") or ""

            if existing:
                education_doc_id = existing["id"]
                cursor.execute("""
                    UPDATE education_documents SET
                        document_type=?, document_type_en=?, document_number=?, completion_date=?,
                        institution_name=?, institution_name_en=?, country=?, country_en=?
                    WHERE id=?
                """, (row["document_type"], doc_type_en, row["document_number"], row["completion_date"],
                      row["institution_name"], inst_name_en, row["country"], row["country_en"], education_doc_id))
                updated += 1
            else:
                cursor.execute("""
                    INSERT INTO education_documents (
                        student_id, document_type, document_type_en, document_number,
                        completion_date, institution_name, institution_name_en, country, country_en
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (row["student_id"], row["document_type"], doc_type_en, row["document_number"],
                      row["completion_date"], row["institution_name"], inst_name_en, row["country"], country_en))
                education_doc_id = cursor.lastrowid
                added += 1

            # --- Foreign / recognition data ---
            cursor.execute(
                "SELECT id FROM foreign_education_docs WHERE education_doc_id = ?",
                (education_doc_id,)
            )
            foreign_exists = cursor.fetchone()

            # Если заполнена хотя бы одна колонка аккредитации/признания
            has_foreign_data = any([
                reference_number,
                reference_institution,
                reference_country,
                reference_issue_date,
                recognition_certificate_number,
                recognition_issuer,
                recognition_date
            ])

            if foreign_exists:
                if has_foreign_data:
                    cursor.execute("""
                        UPDATE foreign_education_docs SET
                            reference_number=?, reference_institution=?, reference_institution_en=?,
                            reference_country=?, reference_country_en=?, reference_issue_date=?,
                            recognition_certificate_number=?, recognition_issuer=?,
                            recognition_issuer_en=?, recognition_date=?
                        WHERE education_doc_id=?
                    """, (
                        reference_number,
                        reference_institution,
                        reference_institution_en,
                        reference_country,
                        reference_country_en,
                        reference_issue_date,
                        recognition_certificate_number,
                        recognition_issuer,
                        recognition_issuer_en,
                        recognition_date,
                        education_doc_id
                    ))
                else:
                    # если все поля пустые - удаляем старую запись
                    cursor.execute(
                        "DELETE FROM foreign_education_docs WHERE education_doc_id=?",
                        (education_doc_id,)
                    )

            else:
                if has_foreign_data:
                    cursor.execute("""
                        INSERT INTO foreign_education_docs (
                            education_doc_id,
                            reference_number,
                            reference_institution,
                            reference_institution_en,
                            reference_country,
                            reference_country_en,
                            reference_issue_date,
                            recognition_certificate_number,
                            recognition_issuer,
                            recognition_issuer_en,
                            recognition_date
                        )
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (
                        education_doc_id,
                        reference_number,
                        reference_institution,
                        reference_institution_en,
                        reference_country,
                        reference_country_en,
                        reference_issue_date,
                        recognition_certificate_number,
                        recognition_issuer,
                        recognition_issuer_en,
                        recognition_date
                    ))

    db.commit()

    log_action(
        session.get('username', 'невідомо'),
        f"імпорт документів про освіту: додано {added}, оновлено {updated}"
    )

    flash(f"Додано записів: {added}, Оновлено записів: {updated}", "success")
    return redirect(url_for('admin.manage_education_documents'))